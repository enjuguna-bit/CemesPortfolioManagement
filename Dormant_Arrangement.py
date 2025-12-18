# Dormant_Arrangement_API.py - API-friendly version
import pandas as pd
import os
from datetime import datetime
import re
import logging
from io import BytesIO
import json
from typing import Dict, List, Tuple, Optional, Any, Union
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

class ProcessingHistory:
    def __init__(self):
        self.history = []
        self.current_index = -1
    
    def add_state(self, df):
        """Add a new state to history"""
        if self.current_index < len(self.history) - 1:
            self.history = self.history[:self.current_index + 1]
        
        self.history.append(df.copy())
        self.current_index += 1
        
        # Limit history size
        if len(self.history) > 10:
            self.history.pop(0)
            self.current_index -= 1
    
    def undo(self):
        """Undo to previous state"""
        if self.current_index > 0:
            self.current_index -= 1
            return self.history[self.current_index].copy()
        return None
    
    def redo(self):
        """Redo to next state"""
        if self.current_index < len(self.history) - 1:
            self.current_index += 1
            return self.history[self.current_index].copy()
        return None

class BranchDataProcessorAPI:
    """API-friendly Branch Data Processor without GUI dependencies."""
    
    # Class-level regex patterns for better performance
    PHONE_PATTERN = re.compile(r'^254\d{9}$')
    
    def __init__(self, log_file: str = 'branch_processor_api.log', enable_history: bool = False):
        self.df = None
        self.branches = []
        # Disable history by default for API usage to save memory
        self.history = ProcessingHistory() if enable_history else None
        
        # Setup logging
        self.setup_logging(log_file)
        
        # Processing options (can be set via API)
        self.options = {
            'remove_duplicates': True,
            'fill_na': True,
            'add_formatting': True,
            'output_format': 'excel'  # 'excel' or 'csv'
        }
    
    def setup_logging(self, log_file: str):
        """Setup logging configuration"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def set_options(self, options: Dict[str, Any]) -> None:
        """Set processing options"""
        valid_options = ['remove_duplicates', 'fill_na', 'add_formatting', 'output_format']
        for key, value in options.items():
            if key in valid_options:
                self.options[key] = value
        self.logger.info(f"Options set: {self.options}")
    
    def load_data(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Load data from bytes (Excel file)"""
        try:
            self.logger.info(f"Loading file: {filename}")
            
            # Read the Excel file from bytes
            if filename.lower().endswith('.csv'):
                # Try different encodings for CSV
                encodings = ['utf-8', 'latin1', 'cp1252', 'iso-8859-1']
                for encoding in encodings:
                    try:
                        content_str = file_content.decode(encoding)
                        self.df = pd.read_csv(BytesIO(content_str.encode()))
                        break
                    except (UnicodeDecodeError, UnicodeError):
                        continue
                else:
                    content_str = file_content.decode('utf-8', errors='ignore')
                    self.df = pd.read_csv(BytesIO(content_str.encode()))
            else:
                # For Excel files
                self.df = pd.read_excel(BytesIO(file_content))
            
            # Extract unique branches
            branch_columns = []
            for col in self.df.columns:
                if 'branch' in col.lower():
                    branch_columns.append(col)
            
            if not branch_columns:
                return {
                    'status': 'error',
                    'message': 'No branch column found in the file!'
                }
            
            # Use the first branch column found
            branch_column = branch_columns[0]
            self.branches = sorted(self.df[branch_column].dropna().unique())
            
            if not self.branches:
                return {
                    'status': 'error',
                    'message': 'No branches found in the file'
                }
            
            self.logger.info(f"Loaded {len(self.df)} records with {len(self.branches)} branches")
            
            return {
                'status': 'success',
                'message': f'Loaded {len(self.df)} records with {len(self.branches)} branches',
                'data_summary': {
                    'total_records': len(self.df),
                    'total_columns': len(self.df.columns),
                    'branches_found': len(self.branches),
                    'branch_names': self.branches[:10],  # First 10 branches
                    'columns': list(self.df.columns.tolist())
                }
            }
                
        except Exception as e:
            error_msg = f"Error loading file: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            return {
                'status': 'error',
                'message': error_msg
            }
    
    def get_branches(self) -> List[str]:
        """Get list of available branches"""
        return self.branches
    
    def get_data_preview(self, rows: int = 50) -> Dict[str, Any]:
        """Get preview of the loaded data"""
        if self.df is None:
            return {
                'status': 'error',
                'message': 'No data loaded'
            }
        
        preview_df = self.df.head(rows).copy()
        
        # Format dates for better display
        for col in preview_df.columns:
            if pd.api.types.is_datetime64_any_dtype(preview_df[col]):
                preview_df[col] = preview_df[col].dt.strftime('%Y-%m-%d')
        
        return {
            'status': 'success',
            'preview': {
                'total_rows': len(self.df),
                'total_columns': len(self.df.columns),
                'unique_branches': len(self.branches),
                'data': preview_df.to_dict(orient='records'),
                'columns': list(preview_df.columns.tolist())
            }
        }
    
    def find_column_by_keywords(self, df, keywords):
        """Find column by matching keywords in column name"""
        for col in df.columns:
            col_lower = str(col).lower()
            for keyword in keywords:
                if keyword in col_lower:
                    return col
        return None
    
    def format_excel_file(self, wb, df):
        """Add formatting to Excel workbook with bold yellow headers"""
        try:
            ws = wb.active
            
            # Define styles
            header_font = Font(name='Arial', size=11, bold=True, color='000000')
            header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            cell_fill_even = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            alignment = Alignment(horizontal='left', vertical='center')
            thin_border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin'))
            
            # Apply formatting to headers (row 1)
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = thin_border
                
                # Auto-adjust column width
                column_letter = get_column_letter(col)
                max_length = 0
                for cell in ws[column_letter]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Apply formatting to data rows with alternating colors
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = alignment
                    cell.border = thin_border
                    
                    # Alternate row colors
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                    else:
                        cell.fill = cell_fill_even
            
            # Add freeze panes for headers
            ws.freeze_panes = 'A2'
            
            self.logger.info("Added formatting to Excel file")
            return wb
            
        except Exception as e:
            self.logger.warning(f"Could not apply Excel formatting: {str(e)}")
            return wb
    
    def normalize_phone_numbers_vectorized(self, df, phone_cols):
        """Normalize phone numbers using vectorized operations for better performance"""
        if not phone_cols:
            return df
        
        df_copy = df.copy()
        
        for phone_col in phone_cols:
            if phone_col in df_copy.columns:
                # Convert to string and fill NaN
                df_copy[phone_col] = df_copy[phone_col].fillna('').astype(str)
                
                # Remove non-digit characters
                df_copy[phone_col] = df_copy[phone_col].str.replace(r'\D', '', regex=True)
                
                # Apply normalization rules using vectorized operations
                mask_07 = df_copy[phone_col].str.startswith('07') & (df_copy[phone_col].str.len() == 10)
                mask_7 = df_copy[phone_col].str.startswith('7') & (df_copy[phone_col].str.len() == 9)
                mask_9digit = (df_copy[phone_col].str.len() == 9) & (~df_copy[phone_col].str.startswith('0'))
                mask_254 = df_copy[phone_col].str.startswith('254') & (df_copy[phone_col].str.len() == 12)
                
                df_copy.loc[mask_07, phone_col] = '254' + df_copy.loc[mask_07, phone_col].str[1:]
                df_copy.loc[mask_7, phone_col] = '254' + df_copy.loc[mask_7, phone_col]
                df_copy.loc[mask_9digit, phone_col] = '254' + df_copy.loc[mask_9digit, phone_col]
                
                # Keep only valid 254xxxxxxxxx format, others set to empty
                invalid_mask = ~mask_254 & ~mask_07 & ~mask_7 & ~mask_9digit
                df_copy.loc[invalid_mask & (df_copy[phone_col].str.len() > 0), phone_col] = ''
                
                # Replace empty strings with 'N/A' if option is checked
                if self.options.get('fill_na', True):
                    df_copy.loc[df_copy[phone_col] == '', phone_col] = 'N/A'
        
        return df_copy
    
    def deduplicate_phone_numbers(self, df, phone_cols):
        """Remove duplicate phone numbers across specified columns"""
        if not phone_cols:
            return df, 0
        
        df_clean = df.copy()
        duplicates_found = 0
        
        for phone_col in phone_cols:
            if phone_col in df_clean.columns:
                # Mark duplicates (keeping first occurrence)
                mask = df_clean.duplicated(subset=[phone_col], keep='first') & (df_clean[phone_col] != '') & (df_clean[phone_col] != 'N/A')
                duplicates_found += mask.sum()
                
                # Remove duplicates
                df_clean = df_clean[~mask]
        
        return df_clean, duplicates_found
    
    def perform_quality_checks(self, df):
        """Perform comprehensive data quality checks"""
        checks = {
            "total_records": len(df),
            "complete_records": int(df.notna().all(axis=1).sum()),
            "records_with_missing_values": int(df.isna().any(axis=1).sum()),
            "duplicate_records": int(df.duplicated().sum()),
            "unique_customers": 0,
            "average_loancount": 0,
        }
        
        # Check for phone number issues
        phone_cols = [col for col in df.columns if any(keyword in col.lower() 
                      for keyword in ['phone', 'mobile', 'number'])]
        
        for phone_col in phone_cols:
            if phone_col in df.columns:
                # Count valid phone numbers using class-level pattern
                valid_phones = df[phone_col].apply(
                    lambda x: bool(self.PHONE_PATTERN.match(str(x))) if pd.notna(x) and str(x).strip() and str(x) != 'N/A' else False
                ).sum()
                
                checks[f"valid_{phone_col}"] = int(valid_phones)
                checks[f"invalid_{phone_col}"] = int(df[phone_col].notna().sum() - valid_phones)
        
        # Count unique customers (by phone if available)
        if phone_cols and phone_cols[0] in df.columns:
            checks["unique_customers"] = int(df[phone_cols[0]].nunique())
        
        # Calculate average LoanCount if column exists
        loancount_col = self.find_column_by_keywords(df, ['loancount'])
        if loancount_col and loancount_col in df.columns:
            checks["average_loancount"] = float(round(df[loancount_col].mean(), 2))
        
        # Check date issues if date column exists
        date_cols = [col for col in df.columns if any(keyword in col.lower() 
                     for keyword in ['date', 'cleared'])]
        
        for date_col in date_cols:
            if date_col in df.columns:
                try:
                    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
                    today = pd.Timestamp.now()
                    
                    # Count future dates
                    future_dates = (df[date_col] > today).sum()
                    checks["records_with_future_dates"] = int(future_dates)
                    
                    # Count very old dates (before 2000)
                    very_old = pd.Timestamp('2000-01-01')
                    checks["records_with_very_old_dates"] = int((df[date_col] < very_old).sum())
                    
                    # Date range
                    valid_dates = df[date_col].dropna()
                    if not valid_dates.empty:
                        checks["earliest_date"] = valid_dates.min().strftime('%Y-%m-%d')
                        checks["latest_date"] = valid_dates.max().strftime('%Y-%m-%d')
                        
                except:
                    pass
        
        return checks
    
    def process_branch(self, branch_name: str) -> Dict[str, Any]:
        """Process data for a single branch"""
        if self.df is None:
            return {
                'status': 'error',
                'message': 'No data loaded. Please load data first.'
            }
        
        if branch_name not in self.branches:
            return {
                'status': 'error',
                'message': f'Branch "{branch_name}" not found. Available branches: {self.branches}'
            }
        
        try:
            self.logger.info(f"Starting data processing for branch: {branch_name}")
            
            # Find branch column
            branch_col = self.find_column_by_keywords(self.df, ['branch'])
            if not branch_col:
                return {
                    'status': 'error',
                    'message': 'No branch column found!'
                }
            
            # Filter data for selected branch
            branch_data = self.df[self.df[branch_col] == branch_name].copy()
            
            if branch_data.empty:
                return {
                    'status': 'error',
                    'message': f'No data found for branch: {branch_name}'
                }
            
            # Save state to history (if enabled)
            if self.history:
                self.history.add_state(branch_data)
            
            # STEP 1: Remove BRANCH column(s)
            branch_columns_to_remove = [col for col in branch_data.columns if 'branch' in col.lower()]
            if branch_columns_to_remove:
                branch_data = branch_data.drop(columns=branch_columns_to_remove)
            
            # STEP 2: Remove DATECREATED column(s)
            datecreated_columns_to_remove = [col for col in branch_data.columns if 'datecreated' in col.lower()]
            if datecreated_columns_to_remove:
                branch_data = branch_data.drop(columns=datecreated_columns_to_remove)
            
            # STEP 3: Find phone columns and normalize
            phone_columns_to_normalize = []
            for col in branch_data.columns:
                if any(keyword in col.lower() for keyword in ['phone', 'mobile', 'number', 'borrowerphone']):
                    phone_columns_to_normalize.append(col)
            
            if phone_columns_to_normalize:
                branch_data = self.normalize_phone_numbers_vectorized(branch_data, phone_columns_to_normalize)
            
            # STEP 4: Remove duplicates if option is checked
            duplicates_removed = 0
            if self.options.get('remove_duplicates', True) and phone_columns_to_normalize:
                branch_data, duplicates_removed = self.deduplicate_phone_numbers(branch_data, phone_columns_to_normalize)
            
            # STEP 5: Find RO/portfolio column
            ro_col = self.find_column_by_keywords(branch_data, ['ro', 'relationship', 'officer', 'portfolio'])
            
            # STEP 6: Find date cleared column
            date_cleared_col = self.find_column_by_keywords(branch_data, ['date', 'cleared', 'dateloancleared'])
            
            # STEP 7: Find first name column
            firstname_col = self.find_column_by_keywords(branch_data, ['firstname', 'first', 'name'])
            
            # Convert date column to datetime if it exists
            if date_cleared_col:
                branch_data[date_cleared_col] = pd.to_datetime(branch_data[date_cleared_col], errors='coerce')
            
            # STEP 8: Sort data
            if ro_col:
                # Get unique officers/portfolios and sort them in ascending order
                officers = sorted(branch_data[ro_col].dropna().unique())
                
                # Create a new DataFrame to store the sorted data
                sorted_data = pd.DataFrame()
                
                for officer in officers:
                    # Filter data for this officer
                    officer_data = branch_data[branch_data[ro_col] == officer].copy()
                    
                    # Sort by date cleared (most recent first) if date column exists
                    if date_cleared_col:
                        officer_data = officer_data.sort_values(by=date_cleared_col, ascending=False)
                    # If no date column, sort by first name
                    elif firstname_col:
                        officer_data = officer_data.sort_values(by=firstname_col, ascending=True)
                    
                    # Add this officer's data to the sorted DataFrame
                    sorted_data = pd.concat([sorted_data, officer_data])
                
                if sorted_data.empty:
                    if date_cleared_col:
                        sorted_data = branch_data.sort_values(by=date_cleared_col, ascending=False)
                    elif firstname_col:
                        sorted_data = branch_data.sort_values(by=firstname_col, ascending=True)
                    else:
                        sorted_data = branch_data
            else:
                if date_cleared_col:
                    sorted_data = branch_data.sort_values(by=date_cleared_col, ascending=False)
                elif firstname_col:
                    sorted_data = branch_data.sort_values(by=firstname_col, ascending=True)
                else:
                    sorted_data = branch_data
            
            # STEP 9: Perform quality checks
            quality_checks = self.perform_quality_checks(sorted_data)
            
            # Prepare response
            response = {
                'status': 'success',
                'branch': branch_name,
                'processing_summary': {
                    'total_records': len(sorted_data),
                    'duplicates_removed': int(duplicates_removed),
                    'phone_columns_normalized': len(phone_columns_to_normalize),
                    'quality_checks': quality_checks
                }
            }
            
            # Add officer distribution if available
            if ro_col:
                officer_counts = sorted_data[ro_col].value_counts().to_dict()
                response['officer_distribution'] = {
                    'total_officers': len(officer_counts),
                    'top_officers': dict(list(officer_counts.items())[:5])  # Top 5 officers
                }
            
            # Add sample data
            response['sample_data'] = sorted_data.head(10).to_dict(orient='records')
            
            # Add processed data for download
            response['processed_data'] = {
                'columns': list(sorted_data.columns.tolist()),
                'row_count': len(sorted_data)
            }
            
            self.logger.info(f"Successfully processed {len(sorted_data)} records for branch: {branch_name}")
            
            return response
            
        except Exception as e:
            error_msg = f"Error processing data: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            return {
                'status': 'error',
                'message': error_msg
            }
    
    def process_all_branches(self) -> Dict[str, Any]:
        """Process all branches at once"""
        if self.df is None:
            return {
                'status': 'error',
                'message': 'No data loaded. Please load data first.'
            }
        
        try:
            results = []
            successful = 0
            failed = []
            
            for branch in self.branches:
                self.logger.info(f"Processing branch: {branch}")
                
                try:
                    # Process each branch
                    branch_result = self.process_branch(branch)
                    
                    if branch_result['status'] == 'success':
                        results.append({
                            'branch': branch,
                            'status': 'success',
                            'record_count': branch_result['processing_summary']['total_records']
                        })
                        successful += 1
                    else:
                        results.append({
                            'branch': branch,
                            'status': 'failed',
                            'error': branch_result['message']
                        })
                        failed.append(branch)
                        
                except Exception as e:
                    results.append({
                        'branch': branch,
                        'status': 'failed',
                        'error': str(e)
                    })
                    failed.append(branch)
            
            return {
                'status': 'success',
                'summary': {
                    'total_branches': len(self.branches),
                    'successful': successful,
                    'failed': len(failed),
                    'failed_branches': failed[:10]  # First 10 failed branches
                },
                'detailed_results': results
            }
            
        except Exception as e:
            error_msg = f"Error processing all branches: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            return {
                'status': 'error',
                'message': error_msg
            }
    
    def generate_report(self) -> Dict[str, Any]:
        """Generate a comprehensive report of the loaded data"""
        if self.df is None:
            return {
                'status': 'error',
                'message': 'No data loaded. Please load data first.'
            }
        
        try:
            # Find branch column
            branch_col = self.find_column_by_keywords(self.df, ['branch'])
            
            # Basic dataset overview
            report = {
                'dataset_overview': {
                    'total_records': len(self.df),
                    'total_columns': len(self.df.columns),
                    'unique_branches': len(self.branches),
                    'file_info': 'Loaded from API'
                }
            }
            
            # Branches information
            if branch_col:
                branch_summary = []
                for i, branch in enumerate(self.branches[:15]):  # First 15 branches
                    branch_count = len(self.df[self.df[branch_col] == branch])
                    branch_summary.append({
                        'rank': i + 1,
                        'branch': branch,
                        'record_count': int(branch_count)
                    })
                
                report['branches_summary'] = branch_summary
            
            # Column information
            column_info = []
            for col in self.df.columns:
                non_null = self.df[col].notna().sum()
                null_count = self.df[col].isna().sum()
                data_type = str(self.df[col].dtype)
                
                # Get sample values for the column
                sample_values = self.df[col].dropna().head(3).tolist()
                
                column_info.append({
                    'column': col,
                    'non_null_count': int(non_null),
                    'null_count': int(null_count),
                    'data_type': data_type,
                    'sample_values': sample_values[:3]  # First 3 non-null values
                })
            
            report['column_information'] = column_info
            
            # Data quality metrics
            quality_metrics = self.perform_quality_checks(self.df)
            report['quality_metrics'] = quality_metrics
            
            return {
                'status': 'success',
                'report': report
            }
            
        except Exception as e:
            error_msg = f"Error generating report: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            return {
                'status': 'error',
                'message': error_msg
            }
    
    def download_processed_data(self, branch_name: str, format: str = 'excel') -> Dict[str, Any]:
        """
        Download processed data for a specific branch
        
        Returns dictionary with file data and metadata
        """
        if self.df is None:
            return {
                'status': 'error',
                'message': 'No data loaded'
            }
        
        if branch_name not in self.branches:
            return {
                'status': 'error',
                'message': f'Branch "{branch_name}" not found'
            }
        
        try:
            # Find branch column
            branch_col = self.find_column_by_keywords(self.df, ['branch'])
            if not branch_col:
                return {
                    'status': 'error',
                    'message': 'No branch column found'
                }
            
            # Filter data for selected branch
            branch_data = self.df[self.df[branch_col] == branch_name].copy()
            
            # Apply basic processing
            # Remove branch columns
            branch_columns_to_remove = [col for col in branch_data.columns if 'branch' in col.lower()]
            if branch_columns_to_remove:
                branch_data = branch_data.drop(columns=branch_columns_to_remove)
            
            # Remove datecreated columns
            datecreated_columns_to_remove = [col for col in branch_data.columns if 'datecreated' in col.lower()]
            if datecreated_columns_to_remove:
                branch_data = branch_data.drop(columns=datecreated_columns_to_remove)
            
            # Normalize phone numbers
            phone_columns_to_normalize = []
            for col in branch_data.columns:
                if any(keyword in col.lower() for keyword in ['phone', 'mobile', 'number', 'borrowerphone']):
                    phone_columns_to_normalize.append(col)
            
            if phone_columns_to_normalize:
                branch_data = self.normalize_phone_numbers_vectorized(branch_data, phone_columns_to_normalize)
            
            # Remove duplicates if option is checked
            if self.options.get('remove_duplicates', True) and phone_columns_to_normalize:
                branch_data, _ = self.deduplicate_phone_numbers(branch_data, phone_columns_to_normalize)
            
            # Sort by date cleared if available
            date_cleared_col = self.find_column_by_keywords(branch_data, ['date', 'cleared', 'dateloancleared'])
            if date_cleared_col:
                branch_data[date_cleared_col] = pd.to_datetime(branch_data[date_cleared_col], errors='coerce')
                branch_data = branch_data.sort_values(by=date_cleared_col, ascending=False)
            
            # Prepare filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            safe_branch_name = branch_name.replace(' ', '_').replace('/', '_')
            
            if format.lower() == 'csv':
                filename = f"{safe_branch_name}_processed_{timestamp}.csv"
                
                # Convert to CSV
                output = BytesIO()
                branch_data.to_csv(output, index=False)
                output.seek(0)
                file_data = output.getvalue()
                mime_type = 'text/csv'
                
            else:  # Excel format
                filename = f"{safe_branch_name}_processed_{timestamp}.xlsx"
                
                # Convert to Excel
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    branch_data.to_excel(writer, index=False, sheet_name='Processed Data')
                
                # Apply formatting if option is checked
                if self.options.get('add_formatting', True):
                    output.seek(0)
                    wb = load_workbook(output)
                    wb = self.format_excel_file(wb, branch_data)
                    
                    # Save formatted workbook back to buffer
                    output = BytesIO()
                    wb.save(output)
                
                output.seek(0)
                file_data = output.getvalue()
                mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
            return {
                'status': 'success',
                'filename': filename,
                'file_data': file_data,
                'mime_type': mime_type,
                'file_size': len(file_data),
                'record_count': len(branch_data),
                'branch': branch_name,
                'timestamp': timestamp
            }
            
        except Exception as e:
            error_msg = f"Error generating download file: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            return {
                'status': 'error',
                'message': error_msg
            }
    
    def batch_download_all_branches(self, format: str = 'excel') -> Dict[str, Any]:
        """Generate a ZIP file with processed data for all branches"""
        try:
            import zipfile
            from io import BytesIO
            
            zip_buffer = BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for branch in self.branches:
                    result = self.download_processed_data(branch, format)
                    
                    if result['status'] == 'success':
                        # Add file to zip
                        zip_file.writestr(result['filename'], result['file_data'])
            
            zip_buffer.seek(0)
            zip_data = zip_buffer.getvalue()
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            zip_filename = f"all_branches_processed_{timestamp}.zip"
            
            return {
                'status': 'success',
                'filename': zip_filename,
                'file_data': zip_data,
                'mime_type': 'application/zip',
                'file_size': len(zip_data),
                'branches_included': len(self.branches),
                'timestamp': timestamp
            }
            
        except Exception as e:
            error_msg = f"Error creating batch download: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            return {
                'status': 'error',
                'message': error_msg
            }

# Flask API integration
from flask import Flask, request, jsonify, send_file
import traceback

app = Flask(__name__)

# Global processor instance (or use application context in production)
processor = BranchDataProcessorAPI()

@app.route('/api/branch/load', methods=['POST'])
def load_branch_data():
    """API endpoint to load branch data"""
    try:
        if 'file' not in request.files:
            return jsonify({
                'status': 'error',
                'message': 'No file provided'
            }), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({
                'status': 'error',
                'message': 'No file selected'
            }), 400
        
        # Read file content
        file_content = file.read()
        
        # Process options if provided
        options = request.form.get('options')
        if options:
            try:
                options_dict = json.loads(options)
                processor.set_options(options_dict)
            except json.JSONDecodeError:
                return jsonify({
                    'status': 'error',
                    'message': 'Invalid options format. Must be valid JSON.'
                }), 400
        
        # Load data
        result = processor.load_data(file_content, file.filename)
        
        if result['status'] == 'error':
            return jsonify(result), 400
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e),
            'traceback': traceback.format_exc()
        }), 500

@app.route('/api/branch/branches', methods=['GET'])
def get_branches():
    """API endpoint to get list of available branches"""
    try:
        branches = processor.get_branches()
        return jsonify({
            'status': 'success',
            'branches': branches,
            'count': len(branches)
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/api/branch/preview', methods=['GET'])
def get_preview():
    """API endpoint to get data preview"""
    try:
        rows = request.args.get('rows', 50, type=int)
        result = processor.get_data_preview(rows)
        
        if result['status'] == 'error':
            return jsonify(result), 400
        
        return jsonify(result)
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/api/branch/process', methods=['POST'])
def process_branch():
    """API endpoint to process a specific branch"""
    try:
        data = request.json or request.form
        
        if not data or 'branch' not in data:
            return jsonify({
                'status': 'error',
                'message': 'Branch name is required'
            }), 400
        
        branch_name = data['branch']
        
        # Process options if provided
        if 'options' in data:
            try:
                if isinstance(data['options'], str):
                    options_dict = json.loads(data['options'])
                else:
                    options_dict = data['options']
                processor.set_options(options_dict)
            except (json.JSONDecodeError, TypeError):
                return jsonify({
                    'status': 'error',
                    'message': 'Invalid options format'
                }), 400
        
        result = processor.process_branch(branch_name)
        
        if result['status'] == 'error':
            return jsonify(result), 400
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e),
            'traceback': traceback.format_exc()
        }), 500

@app.route('/api/branch/process-all', methods=['POST'])
def process_all_branches():
    """API endpoint to process all branches"""
    try:
        # Process options if provided
        data = request.json or request.form
        if data and 'options' in data:
            try:
                if isinstance(data['options'], str):
                    options_dict = json.loads(data['options'])
                else:
                    options_dict = data['options']
                processor.set_options(options_dict)
            except (json.JSONDecodeError, TypeError):
                return jsonify({
                    'status': 'error',
                    'message': 'Invalid options format'
                }), 400
        
        result = processor.process_all_branches()
        
        if result['status'] == 'error':
            return jsonify(result), 400
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e),
            'traceback': traceback.format_exc()
        }), 500

@app.route('/api/branch/report', methods=['GET'])
def get_report():
    """API endpoint to generate a report"""
    try:
        result = processor.generate_report()
        
        if result['status'] == 'error':
            return jsonify(result), 400
        
        return jsonify(result)
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/api/branch/download', methods=['POST'])
def download_branch():
    """API endpoint to download processed data for a branch"""
    try:
        data = request.json or request.form
        
        if not data or 'branch' not in data:
            return jsonify({
                'status': 'error',
                'message': 'Branch name is required'
            }), 400
        
        branch_name = data['branch']
        format = data.get('format', 'excel')
        
        result = processor.download_processed_data(branch_name, format)
        
        if result['status'] == 'error':
            return jsonify(result), 400
        
        # Return file as download
        return send_file(
            BytesIO(result['file_data']),
            as_attachment=True,
            download_name=result['filename'],
            mimetype=result['mime_type']
        )
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e),
            'traceback': traceback.format_exc()
        }), 500

@app.route('/api/branch/download-all', methods=['GET'])
def download_all_branches():
    """API endpoint to download all processed branches as ZIP"""
    try:
        format = request.args.get('format', 'excel')
        
        result = processor.batch_download_all_branches(format)
        
        if result['status'] == 'error':
            return jsonify(result), 400
        
        # Return file as download
        return send_file(
            BytesIO(result['file_data']),
            as_attachment=True,
            download_name=result['filename'],
            mimetype=result['mime_type']
        )
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': str(e),
            'traceback': traceback.format_exc()
        }), 500

@app.route('/api/branch/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'service': 'Branch Data Processor API',
        'version': '1.0.0',
        'timestamp': datetime.now().isoformat()
    })

# Command line usage example
if __name__ == "__main__":
    # Example of how to use the processor directly
    import sys
    
    if len(sys.argv) > 1:
        # Read file from command line
        file_path = sys.argv[1]
        
        with open(file_path, 'rb') as f:
            file_content = f.read()
        
        processor = BranchDataProcessorAPI()
        result = processor.load_data(file_content, os.path.basename(file_path))
        
        print(json.dumps(result, indent=2))
    else:
        print("Starting Branch Data Processor API...")
        print("Usage: python dormant_arrangement_api.py <file_path>")
        print("Or start the Flask API server")
        
        # Start Flask server for API
        app.run(host='0.0.0.0', port=5001, debug=True)