# loan_report_generator.py - Refactored for API without tkinter

import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
import numpy as np
from typing import List, Dict, Any, Optional, Union
import warnings
import io
import tempfile

warnings.filterwarnings('ignore')

class LoanReportGeneratorAPI:
    """API version of Loan Report Generator without tkinter"""
    
    def __init__(self):
        self.df = None
        self.final_df = None
        self.groups = []
        self.grand_totals = {}
        self.output_file = None
        self.temp_files = []
    
    def format_phone_number(self, phone: Any) -> str:
        """Format phone number to standard Kenyan format"""
        if pd.isna(phone):
            return ""
        
        phone_str = str(phone).strip()
        
        # Remove any non-digit characters except plus sign
        phone_digits = ''.join(filter(lambda x: x.isdigit() or x == '+', phone_str))
        
        # Remove plus sign for processing
        if phone_digits.startswith('+'):
            phone_digits = phone_digits[1:]
        
        # Format based on length and prefix
        if len(phone_digits) == 12 and phone_digits.startswith('254'):
            return f"+254 {phone_digits[3:6]} {phone_digits[6:9]} {phone_digits[9:]}"
        elif len(phone_digits) == 9:
            return f"+254 {phone_digits[0:3]} {phone_digits[3:6]} {phone_digits[6:]}"
        elif len(phone_digits) == 10 and phone_digits.startswith('0'):
            return f"+254 {phone_digits[1:4]} {phone_digits[4:7]} {phone_digits[7:]}"
        else:
            return phone_str  # Return original if can't format properly

    def load_and_prepare_data(self, file_input: Union[str, io.BytesIO]) -> Dict:
        """Load CSV/Excel file and prepare data for processing"""
        try:
            # Determine file type and load accordingly
            if isinstance(file_input, io.BytesIO):
                # Reset stream position
                file_input.seek(0)
                
                # Try different formats
                try:
                    # Try Excel first
                    self.df = pd.read_excel(file_input)
                except:
                    # Reset and try CSV
                    file_input.seek(0)
                    self.df = pd.read_csv(file_input)
            else:
                # File path
                if not os.path.exists(file_input):
                    return {
                        'status': 'error',
                        'message': f"The file '{file_input}' was not found."
                    }
                
                if file_input.lower().endswith('.csv'):
                    # Try different encodings for CSV
                    encodings = ['utf-8', 'latin-1', 'iso-8859-1']
                    for encoding in encodings:
                        try:
                            self.df = pd.read_csv(file_input, encoding=encoding)
                            break
                        except UnicodeDecodeError:
                            continue
                else:
                    # Excel file
                    self.df = pd.read_excel(file_input)
            
            if self.df is None:
                return {
                    'status': 'error',
                    'message': "Could not read the file with supported formats/encodings."
                }
            
            # Clean column names
            self.df.columns = self.df.columns.str.strip().str.replace('"', '')
            
            # Validate required columns
            required_columns = ['FullNames', 'FieldOfficer']
            missing_columns = [col for col in required_columns if col not in self.df.columns]
            
            if missing_columns:
                return {
                    'status': 'error',
                    'message': f"Missing required columns: {missing_columns}\nAvailable columns: {list(self.df.columns)}"
                }
            
            return {
                'status': 'success',
                'message': f'Successfully loaded {len(self.df)} records',
                'record_count': len(self.df),
                'columns': list(self.df.columns)
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Failed to load file: {str(e)}'
            }

    def process_data(self) -> Dict:
        """Process and transform the data"""
        try:
            # ==========================================
            # 1. COLUMN SELECTION AND REORDERING
            # ==========================================
            
            # Define columns to keep and their order
            column_config = {
                'FullNames': 'Client Name',
                'PhoneNumber': 'Phone Number',
                'InstallmentNo': 'Installment No',
                'Amount Due': 'Amount Due',
                'Arrears': 'Arrears',
                'AmountPaid': 'Amount Paid',
                'LoanBalance': 'Loan Balance',
                'FieldOfficer': 'Field Officer'
            }
            
            # Create new DataFrame with only required columns
            processed_df = pd.DataFrame()
            
            for original_col, new_name in column_config.items():
                if original_col in self.df.columns:
                    processed_df[new_name] = self.df[original_col]
                else:
                    # If column doesn't exist, create empty column
                    processed_df[new_name] = None
            
            # ==========================================
            # 2. DATA TRANSFORMATIONS
            # ==========================================
            
            # Format phone numbers
            if 'Phone Number' in processed_df.columns:
                processed_df['Phone Number'] = processed_df['Phone Number'].apply(self.format_phone_number)
            
            # Convert and clean numeric columns
            numeric_columns = ['Amount Due', 'Arrears', 'Amount Paid', 'Loan Balance']
            for col in numeric_columns:
                if col in processed_df.columns:
                    # Convert to string, remove non-numeric characters, then to float
                    processed_df[col] = (
                        processed_df[col]
                        .astype(str)
                        .str.replace(r'[^\d\.\-]', '', regex=True)
                        .replace('', '0')
                    )
                    processed_df[col] = pd.to_numeric(processed_df[col], errors='coerce').fillna(0)
            
            # Convert Installment No to integer
            if 'Installment No' in processed_df.columns:
                processed_df['Installment No'] = pd.to_numeric(
                    processed_df['Installment No'], errors='coerce'
                ).fillna(0).astype(int)
            
            # ==========================================
            # 3. SORTING AND GROUPING
            # ==========================================
            
            if 'Field Officer' in processed_df.columns and 'Installment No' in processed_df.columns:
                # Replace empty field officer names
                processed_df['Field Officer'] = processed_df['Field Officer'].fillna('Unknown')
                processed_df['Field Officer'] = processed_df['Field Officer'].astype(str).str.strip()
                
                # Make sure Client Name exists for sorting
                if 'Client Name' not in processed_df.columns:
                    return {
                        'status': 'error',
                        'message': "'Client Name' column not found after mapping!"
                    }
                
                # Sort data
                processed_df = processed_df.sort_values(
                    ['Field Officer', 'Installment No', 'Client Name'],
                    ascending=[True, True, True]
                )
            
            self.df = processed_df
            
            return {
                'status': 'success',
                'message': f'Data processed successfully. Shape: {self.df.shape}',
                'columns': list(self.df.columns)
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Error in data processing: {str(e)}'
            }

    def generate_report_structure(self) -> Dict:
        """Generate the final report structure with groups and totals"""
        try:
            grouping_column = 'Field Officer'
            
            if grouping_column not in self.df.columns:
                return {
                    'status': 'error',
                    'message': f"Grouping column '{grouping_column}' not found in DataFrame"
                }
            
            final_rows = []
            
            # Get unique groups
            self.groups = sorted(self.df[grouping_column].unique())
            
            # Initialize grand totals
            numeric_columns = ['Amount Due', 'Arrears', 'Amount Paid', 'Loan Balance']
            numeric_columns = [col for col in numeric_columns if col in self.df.columns]
            self.grand_totals = {col: 0 for col in numeric_columns}
            
            # Process each group
            for group_idx, group_name in enumerate(self.groups, 1):
                group_data = self.df[self.df[grouping_column] == group_name].copy()
                
                # A. Group Header
                header_row = {col: '' for col in self.df.columns}
                header_row['Client Name'] = f"--- {group_name.upper()} ({len(group_data)} clients) ---"
                final_rows.append(header_row)
                
                # B. Group Data
                for _, row in group_data.iterrows():
                    row_dict = {}
                    for col in self.df.columns:
                        value = row[col]
                        if col in numeric_columns and pd.notna(value):
                            row_dict[col] = float(value)
                        else:
                            row_dict[col] = value
                    final_rows.append(row_dict)
                
                # C. Group Subtotal
                subtotal_row = {col: '' for col in self.df.columns}
                subtotal_row['Client Name'] = f"Subtotal {group_name}"
                
                group_totals = {}
                for col in numeric_columns:
                    subtotal = group_data[col].sum()
                    subtotal_row[col] = subtotal
                    group_totals[col] = subtotal
                    self.grand_totals[col] += subtotal
                
                if 'Installment No' in group_data.columns:
                    subtotal_row['Installment No'] = group_data['Installment No'].sum()
                
                final_rows.append(subtotal_row)
                
                # D. Group Summary (compact format)
                summary_row = {col: '' for col in self.df.columns}
                summary_values = [f"{group_totals.get(col, 0):,.2f}" for col in numeric_columns]
                summary_row['Client Name'] = " | ".join(summary_values)
                final_rows.append(summary_row)
                
                # E. Separator (except for last group)
                if group_idx < len(self.groups):
                    final_rows.append({col: '' for col in self.df.columns})
            
            # F. Grand Total
            grand_total_row = {col: '' for col in self.df.columns}
            grand_total_row['Client Name'] = "GRAND TOTAL"
            
            for col in numeric_columns:
                grand_total_row[col] = self.grand_totals[col]
            
            final_rows.append(grand_total_row)
            
            # G. Final Summary
            final_summary_row = {col: '' for col in self.df.columns}
            final_summary_values = [f"{self.grand_totals.get(col, 0):,.2f}" for col in numeric_columns]
            final_summary_row['Client Name'] = " | ".join(final_summary_values)
            final_rows.append(final_summary_row)
            
            # Add summary statistics
            self.add_summary_statistics(final_rows)
            
            # Create final DataFrame
            self.final_df = pd.DataFrame(final_rows)
            
            return {
                'status': 'success',
                'message': f'Report structure generated with {len(self.final_df)} rows',
                'group_count': len(self.groups),
                'total_rows': len(self.final_df),
                'grand_totals': self.grand_totals
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Error generating report structure: {str(e)}'
            }

    def add_summary_statistics(self, final_rows: List[Dict]):
        """Add summary statistics to the report"""
        try:
            # Calculate statistics
            total_clients = len(self.df)
            total_groups = len(self.groups)
            
            summary_stats = []
            
            # Clients with arrears
            if 'Arrears' in self.df.columns:
                clients_with_arrears = len(self.df[self.df['Arrears'] > 0])
                arrears_percentage = (clients_with_arrears / total_clients * 100) if total_clients > 0 else 0
                
                summary_stats.append(f"Total Clients: {total_clients:,}")
                summary_stats.append(f"Field Officers: {total_groups}")
                summary_stats.append(f"Clients with Arrears: {clients_with_arrears:,} ({arrears_percentage:.1f}%)")
                summary_stats.append(f"Total Arrears: KES {self.grand_totals.get('Arrears', 0):,.2f}")
                summary_stats.append(f"Total Amount Due: KES {self.grand_totals.get('Amount Due', 0):,.2f}")
            
            # Add statistics as a final row
            if summary_stats:
                stats_row = {col: '' for col in self.final_df.columns} if hasattr(self, 'final_df') else {}
                stats_row['Client Name'] = "SUMMARY: " + " | ".join(summary_stats)
                final_rows.append(stats_row)
                
        except Exception as e:
            print(f"Error adding summary statistics: {e}")

    def apply_excel_formatting(self, workbook_path: str) -> Dict:
        """Apply advanced formatting to the Excel file"""
        try:
            workbook = load_workbook(workbook_path)
            worksheet = workbook.active
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            group_header_font = Font(bold=True, color="1F4E78", size=11)
            subtotal_font = Font(bold=True, color="C65911", size=10)
            total_font = Font(bold=True, color="FF0000", size=11)
            arrears_font = Font(color="FF0000", bold=True)  # Red for arrears > 0
            currency_format = '#,##0.00'
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Find column indices
            column_indices = {}
            for idx, cell in enumerate(worksheet[1], 1):
                column_indices[cell.value] = idx
            
            # Apply formatting to all rows
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                # Apply borders to all cells
                for cell in row:
                    cell.border = thin_border
                
                # Header row formatting
                if row[0].row == 1:
                    for cell in row:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Group header rows (--- FIELD OFFICER ---)
                elif '---' in str(row[0].value):
                    for cell in row:
                        cell.font = group_header_font
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Subtotal rows
                elif str(row[0].value).startswith('Subtotal') or str(row[0].value).startswith('Total'):
                    for cell in row:
                        cell.font = subtotal_font
                
                # Grand total row
                elif str(row[0].value) == 'GRAND TOTAL':
                    for cell in row:
                        cell.font = total_font
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                # Summary row (totals line)
                elif '|' in str(row[0].value):
                    for cell in row:
                        cell.font = Font(bold=True, italic=True)
                        cell.alignment = Alignment(horizontal='center')
                
                # Arrears column formatting
                if 'Arrears' in column_indices:
                    arrears_cell = row[column_indices['Arrears'] - 1]
                    try:
                        arrears_value = float(str(arrears_cell.value)) if arrears_cell.value not in [None, ''] else 0
                        if arrears_value > 0:
                            arrears_cell.font = arrears_font
                    except (ValueError, TypeError):
                        pass
                
                # Apply currency formatting to numeric columns
                numeric_columns = ['Amount Due', 'Arrears', 'Amount Paid', 'Loan Balance']
                for col_name in numeric_columns:
                    if col_name in column_indices:
                        cell = row[column_indices[col_name] - 1]
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = currency_format
            
            # Auto-adjust column widths with better logic
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                # Find maximum length in column
                for cell in column:
                    try:
                        if cell.value:
                            # Count characters for display
                            cell_length = len(str(cell.value))
                            # Add extra for currency formatting
                            if cell.number_format == currency_format:
                                cell_length += 3
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # Set width with limits
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
            
            # Add filters
            worksheet.auto_filter.ref = worksheet.dimensions
            
            workbook.save(workbook_path)
            
            return {
                'status': 'success',
                'message': 'Excel formatting applied successfully'
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Error applying Excel formatting: {str(e)}'
            }

    def generate_report(self, file_input: Union[str, io.BytesIO], 
                       output_path: Optional[str] = None,
                       output_format: str = 'xlsx') -> Dict:
        """
        Main function to generate arranged report
        
        Args:
            file_input: File path or BytesIO object
            output_path: Optional output path for the report
            output_format: Output format - 'xlsx' or 'csv'
        
        Returns:
            Dictionary with results
        """
        try:
            # Step 1: Load data
            load_result = self.load_and_prepare_data(file_input)
            if load_result['status'] == 'error':
                return load_result
            
            # Step 2: Process data
            process_result = self.process_data()
            if process_result['status'] == 'error':
                return process_result
            
            # Step 3: Generate report structure
            report_result = self.generate_report_structure()
            if report_result['status'] == 'error':
                return report_result
            
            # Step 4: Create output file
            if output_path is None:
                # Create temporary file
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                temp_file = tempfile.NamedTemporaryFile(
                    suffix=f'.{output_format}', 
                    prefix=f'Loan_Report_{timestamp}_',
                    delete=False
                )
                output_path = temp_file.name
                temp_file.close()
                self.temp_files.append(output_path)
            
            self.output_file = output_path
            
            # Step 5: Save in requested format
            if output_format.lower() == 'xlsx':
                # Save to Excel
                self.final_df.to_excel(output_path, index=False, sheet_name='Loan Report')
                
                # Apply Excel formatting
                format_result = self.apply_excel_formatting(output_path)
                if format_result['status'] == 'error':
                    return format_result
                
            elif output_format.lower() == 'csv':
                # Save to CSV
                self.final_df.to_csv(output_path, index=False)
            else:
                return {
                    'status': 'error',
                    'message': f'Unsupported output format: {output_format}'
                }
            
            # Step 6: Prepare summary statistics for response
            if 'Arrears' in self.df.columns:
                total_clients = len(self.df)
                clients_with_arrears = len(self.df[self.df['Arrears'] > 0])
                arrears_percentage = (clients_with_arrears / total_clients * 100) if total_clients > 0 else 0
                
                summary = {
                    'total_clients': total_clients,
                    'field_officers': len(self.groups),
                    'clients_with_arrears': clients_with_arrears,
                    'arrears_percentage': round(arrears_percentage, 2),
                    'total_arrears': float(self.grand_totals.get('Arrears', 0)),
                    'total_amount_due': float(self.grand_totals.get('Amount Due', 0)),
                    'total_loan_balance': float(self.grand_totals.get('Loan Balance', 0))
                }
            else:
                summary = {
                    'total_clients': len(self.df),
                    'field_officers': len(self.groups)
                }
            
            # Step 7: Return success response
            return {
                'status': 'success',
                'message': 'Report generated successfully',
                'output_file': output_path,
                'output_filename': os.path.basename(output_path),
                'summary': summary,
                'file_size': os.path.getsize(output_path) if os.path.exists(output_path) else 0,
                'metadata': {
                    'record_count': load_result['record_count'],
                    'group_count': len(self.groups),
                    'output_format': output_format,
                    'generated_at': datetime.now().isoformat()
                }
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Error generating report: {str(e)}'
            }
    
    def cleanup(self):
        """Clean up temporary files"""
        for temp_file in self.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except:
                pass


# Helper function for API usage
def generate_loan_report(file_input: Union[str, io.BytesIO], 
                        output_path: Optional[str] = None,
                        output_format: str = 'xlsx') -> Dict:
    """
    Convenience function for API usage
    
    Args:
        file_input: File path or BytesIO object
        output_path: Optional output path for the report
        output_format: Output format - 'xlsx' or 'csv'
    
    Returns:
        Dictionary with analysis results
    """
    generator = LoanReportGeneratorAPI()
    result = generator.generate_report(file_input, output_path, output_format)
    
    # Clean up temporary files on error
    if result['status'] == 'error':
        generator.cleanup()
    
    return result


# For direct script execution
if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Generate loan report from CSV/Excel file')
    parser.add_argument('input_file', help='Path to input CSV or Excel file')
    parser.add_argument('-o', '--output', help='Output file path (optional)')
    parser.add_argument('-f', '--format', choices=['xlsx', 'csv'], default='xlsx', 
                       help='Output format (default: xlsx)')
    
    args = parser.parse_args()
    
    result = generate_loan_report(args.input_file, args.output, args.format)
    
    if result['status'] == 'success':
        print(f"✓ Report generated successfully!")
        print(f"Output file: {result['output_file']}")
        print(f"Summary: {result['summary']}")
    else:
        print(f"✗ Error: {result.get('message', 'Unknown error')}")