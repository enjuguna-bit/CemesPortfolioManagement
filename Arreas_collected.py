# Arrears_collected_api.py - API-friendly version
import pandas as pd
from datetime import datetime
from typing import Optional, Dict, Tuple, List
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from io import BytesIO
import tempfile
import os

class ArrearsProcessorAPI:
    """API-friendly Arrears Processor without GUI dependencies."""
    
    # Define constants - use cleaned column names (without spaces)
    VALID_BUCKETS = ['1-15', '16-30', '31-180']
    REQUIRED_COLUMNS_SOD = ['LoanId', 'SalesRep', 'ArrearsAmount', 'DaysInArrears']
    REQUIRED_COLUMNS_CUR = ['LoanId', 'ArrearsAmount']
    
    def __init__(self):
        self.officer_targets = {}
        
    @staticmethod
    def get_bucket(days: int) -> str:
        """Categorize arrears into age buckets."""
        if 1 <= days <= 15:
            return '1-15'
        elif 16 <= days <= 30:
            return '16-30'
        elif 31 <= days <= 180:
            return '31-180'
        return 'Other'
    
    def normalize_officer_names(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and normalize officer names."""
        if 'SalesRep' in df.columns:
            # Remove leading/trailing whitespace
            df['SalesRep'] = df['SalesRep'].astype(str).str.strip()
            
            # Replace common variations
            name_corrections = {
                'Brian Wanj:': 'Brian Wanjau',
                'Brian Wanjau:': 'Brian Wanjau',
                'Brian W.': 'Brian Wanjau',
                # Add more corrections as needed
            }
            
            for wrong_name, correct_name in name_corrections.items():
                df['SalesRep'] = df['SalesRep'].str.replace(wrong_name, correct_name, regex=False)
            
            # Capitalize first letter of each word
            df['SalesRep'] = df['SalesRep'].str.title()
            
        return df
    
    def parse_targets_from_json(self, targets_json: Dict[str, float]) -> Dict[str, float]:
        """Parse officer targets from JSON dictionary."""
        if not targets_json:
            return {}
        
        processed_targets = {}
        for officer, target in targets_json.items():
            try:
                if isinstance(target, (int, float)):
                    processed_targets[officer] = round(float(target), 2)
                elif isinstance(target, str):
                    target = target.replace(',', '').strip()
                    if target == "":
                        processed_targets[officer] = 0.0
                    else:
                        processed_targets[officer] = round(float(target), 2)
                else:
                    processed_targets[officer] = 0.0
            except (ValueError, TypeError):
                processed_targets[officer] = 0.0
        
        return processed_targets
    
    def validate_dataframes(self, df_sod: pd.DataFrame, df_cur: pd.DataFrame) -> Tuple[bool, str]:
        """Validate that required columns exist in the dataframes."""
        # Check using the cleaned column names (without spaces)
        missing_sod = [col for col in self.REQUIRED_COLUMNS_SOD if col not in df_sod.columns]
        missing_cur = [col for col in self.REQUIRED_COLUMNS_CUR if col not in df_cur.columns]
        
        if missing_sod:
            # Show the original expected column names for user understanding
            original_names = {
                'LoanId': 'LoanId or Loan ID',
                'SalesRep': 'SalesRep or Sales Rep',
                'ArrearsAmount': 'Arrears Amount',
                'DaysInArrears': 'DaysInArrears or Days In Arrears'
            }
            missing_readable = [original_names.get(col, col) for col in missing_sod]
            return False, f"Start-of-Day file missing columns: {', '.join(missing_readable)}"
        if missing_cur:
            original_names = {
                'LoanId': 'LoanId or Loan ID',
                'ArrearsAmount': 'Arrears Amount'
            }
            missing_readable = [original_names.get(col, col) for col in missing_cur]
            return False, f"Current file missing columns: {', '.join(missing_readable)}"
        
        return True, "Validation successful"
    
    def load_and_clean_data(self, file_content: bytes, filename: str) -> pd.DataFrame:
        """Load data from bytes with error handling and cleaning."""
        try:
            # Determine file type from filename
            if filename.lower().endswith('.csv'):
                # Try different encodings for CSV files
                encodings = ['utf-8', 'latin1', 'cp1252', 'iso-8859-1']
                
                for encoding in encodings:
                    try:
                        # Convert bytes to string with appropriate encoding
                        content_str = file_content.decode(encoding)
                        df = pd.read_csv(io.StringIO(content_str))
                        break
                    except (UnicodeDecodeError, UnicodeError):
                        continue
                else:
                    # If all encodings fail, try with default encoding
                    content_str = file_content.decode('utf-8', errors='ignore')
                    df = pd.read_csv(io.StringIO(content_str))
            else:
                # For Excel files
                df = pd.read_excel(io.BytesIO(file_content))
            
            # Clean column names (remove extra spaces, lowercase, replace spaces with nothing)
            df.columns = df.columns.str.strip().str.title().str.replace(' ', '')
            
            # Map common column name variations - AFTER cleaning
            column_mapping = {
                'Loanid': 'LoanId',
                'Loanid': 'LoanId',
                'LoanId': 'LoanId',
                'Loan_Id': 'LoanId',
                
                'Salesrep': 'SalesRep',
                'Salesrep': 'SalesRep',
                'SalesRep': 'SalesRep',
                'Sales_Rep': 'SalesRep',
                'Officer': 'SalesRep',
                'Agent': 'SalesRep',
                
                'Arrearsamount': 'ArrearsAmount',
                'Arrearsamount': 'ArrearsAmount',
                'ArrearsAmount': 'ArrearsAmount',
                'Arrears_Amount': 'ArrearsAmount',
                'Arrears': 'ArrearsAmount',
                'Amount': 'ArrearsAmount',
                'Balance': 'ArrearsAmount',
                
                'Daysinarrears': 'DaysInArrears',
                'Daysinarrears': 'DaysInArrears',
                'DaysInArrears': 'DaysInArrears',
                'Days_In_Arrears': 'DaysInArrears',
                'Days': 'DaysInArrears',
                'Age': 'DaysInArrears',
                'Daysinarr': 'DaysInArrears',
            }
            
            # Apply column mapping
            df.rename(columns=column_mapping, inplace=True)
            
            return df
            
        except Exception as e:
            raise Exception(f"Error loading {filename}: {str(e)}")
    
    def process_data(self, sod_content: bytes, sod_filename: str, 
                    cur_content: bytes, cur_filename: str) -> Optional[tuple]:
        """Process the arrears data and return pivot table."""
        try:
            # Load and clean data
            df_sod = self.load_and_clean_data(sod_content, sod_filename)
            df_cur = self.load_and_clean_data(cur_content, cur_filename)
            
            # Normalize officer names
            df_sod = self.normalize_officer_names(df_sod)
            
            # Validate columns
            is_valid, message = self.validate_dataframes(df_sod, df_cur)
            if not is_valid:
                raise ValueError(message)
            
            # Rename columns for clarity
            df_sod = df_sod.rename(columns={
                'ArrearsAmount': 'Arrears_SOD', 
                'DaysInArrears': 'Age_SOD'
            })
            df_cur = df_cur.rename(columns={'ArrearsAmount': 'Arrears_CUR'})
            
            # Convert numeric columns, coerce errors to NaN
            df_sod['Arrears_SOD'] = pd.to_numeric(df_sod['Arrears_SOD'], errors='coerce')
            df_sod['Age_SOD'] = pd.to_numeric(df_sod['Age_SOD'], errors='coerce')
            df_cur['Arrears_CUR'] = pd.to_numeric(df_cur['Arrears_CUR'], errors='coerce')
            
            # Drop rows with NaN in critical columns
            df_sod = df_sod.dropna(subset=['LoanId', 'SalesRep', 'Arrears_SOD'])
            df_cur = df_cur.dropna(subset=['LoanId', 'Arrears_CUR'])
            
            # Get unique officers
            unique_officers = sorted(df_sod['SalesRep'].unique().tolist())
            
            # Merge dataframes
            df_merged = pd.merge(
                df_sod[['LoanId', 'SalesRep', 'Arrears_SOD', 'Age_SOD']],
                df_cur[['LoanId', 'Arrears_CUR']],
                on='LoanId',
                how='left',
                validate='one_to_one'
            )
            
            # Handle missing values (paid-off loans)
            df_merged['Arrears_CUR'] = df_merged['Arrears_CUR'].fillna(0)
            
            # Calculate collections
            df_merged['Collected'] = df_merged['Arrears_SOD'] - df_merged['Arrears_CUR']
            
            # Filter and categorize
            df_collected = df_merged[df_merged['Collected'] > 0].copy()
            df_collected['Bucket'] = df_collected['Age_SOD'].apply(self.get_bucket)
            
            # Filter for valid buckets only
            df_collected = df_collected[df_collected['Bucket'].isin(self.VALID_BUCKETS)]
            
            return df_collected, df_merged, unique_officers
            
        except Exception as e:
            raise Exception(f"Error processing data: {str(e)}")
    
    def create_formatted_table(self, df_collected: pd.DataFrame, 
                              officers: List[str],
                              officer_targets: Optional[Dict[str, float]] = None) -> pd.DataFrame:
        """Create formatted pivot table with targets and remaining calculations."""
        # Create pivot table
        pivot = pd.pivot_table(
            df_collected,
            values='Collected',
            index='SalesRep',
            columns='Bucket',
            aggfunc='sum',
            fill_value=0,
            margins=False
        )
        
        # Ensure all officers are represented (even if they collected 0)
        for officer in officers:
            if officer not in pivot.index:
                # Add row with zeros for this officer
                new_row = pd.Series({bucket: 0.0 for bucket in self.VALID_BUCKETS}, name=officer)
                pivot = pd.concat([pivot, pd.DataFrame([new_row])])
        
        # Ensure all bucket columns exist
        for bucket in self.VALID_BUCKETS:
            if bucket not in pivot.columns:
                pivot[bucket] = 0.0
        
        # Reorder columns
        pivot = pivot[self.VALID_BUCKETS]
        
        # Add Grand Total column
        pivot['Grand Total'] = pivot.sum(axis=1)
        
        # Sort by Grand Total descending
        pivot = pivot.sort_values('Grand Total', ascending=False)
        
        # Get list of officers from pivot (including those with 0 collections)
        all_officers = pivot.index.tolist()
        
        # Use provided officer targets
        if officer_targets is not None:
            self.officer_targets = self.parse_targets_from_json(officer_targets)
        
        # Add Targets column if targets were collected
        if self.officer_targets:
            # Ensure all officers have targets
            for officer in pivot.index:
                if officer not in self.officer_targets:
                    self.officer_targets[officer] = 0.0
            
            pivot['Target'] = pivot.index.map(self.officer_targets)
            
            # Calculate Remaining (Target - Grand Total)
            pivot['Remaining'] = pivot['Target'] - pivot['Grand Total']
            
            # Add achievement percentage (only if target > 0)
            pivot['Achievement %'] = np.where(
                pivot['Target'] > 0,
                (pivot['Grand Total'] / pivot['Target'] * 100),
                0
            )
        
        # Add Grand Total row
        grand_row = pd.DataFrame(pivot.sum(axis=0)).T
        grand_row.index = ['GRAND TOTAL']
        
        # Calculate totals for Target, Remaining, and Achievement %
        if 'Target' in pivot.columns:
            grand_row['Target'] = pivot['Target'].sum()
            grand_row['Remaining'] = grand_row['Target'] - grand_row['Grand Total']
            if grand_row['Target'].iloc[0] > 0:
                grand_row['Achievement %'] = (grand_row['Grand Total'] / grand_row['Target'] * 100)
            else:
                grand_row['Achievement %'] = 0
        
        final_df = pd.concat([pivot, grand_row])
        
        # Round numeric values to 2 decimal places
        numeric_cols = final_df.select_dtypes(include=[np.number]).columns
        final_df[numeric_cols] = final_df[numeric_cols].round(2)
        
        return final_df
    
    def format_excel_report(self, wb):
        """Apply professional formatting to the Excel workbook."""
        # Format Summary Report sheet
        if 'Summary Report' in wb.sheetnames:
            ws = wb['Summary Report']
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            total_font = Font(bold=True, color="FFFFFF")
            total_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
            center_align = Alignment(horizontal='center', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format headers
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            # Format data cells
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    
                    # Right align for numeric columns (skip first column - officer names)
                    if col > 1 and cell.value:
                        try:
                            # Check if it's a number
                            float(str(cell.value).replace(',', '').replace('%', ''))
                            cell.alignment = right_align
                        except:
                            pass
            
            # Format GRAND TOTAL row
            grand_total_row = ws.max_row
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=grand_total_row, column=col)
                cell.font = total_font
                cell.fill = total_fill
                cell.border = thin_border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        return wb
    
    def create_excel_report(self, df_collected: pd.DataFrame, df_merged: pd.DataFrame, 
                           final_df: pd.DataFrame) -> BytesIO:
        """Create formatted Excel report with multiple sheets in memory."""
        # Create in-memory Excel file
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Summary Report
            summary_sheet = 'Summary Report'
            final_df.to_excel(writer, sheet_name=summary_sheet)
            
            # Sheet 2: Detailed Data
            detailed_sheet = 'Detailed Collections'
            df_collected.to_excel(writer, sheet_name=detailed_sheet, index=False)
            
            # Sheet 3: Raw Data
            raw_sheet = 'Raw Data'
            df_merged.to_excel(writer, sheet_name=raw_sheet, index=False)
            
            # Sheet 4: Statistics
            total_collected = df_collected['Collected'].sum() if not df_collected.empty else 0
            avg_collected = df_collected['Collected'].mean() if not df_collected.empty else 0
            officer_count = len(df_collected['SalesRep'].unique()) if not df_collected.empty else 0
            
            stats_data = {
                'Metric': ['Total Loans Collected', 'Total Amount Collected', 'Average per Loan',
                          'Number of Officers', 'Total Loans Processed'],
                'Value': [
                    len(df_collected),
                    f"{total_collected:,.2f}",
                    f"{avg_collected:,.2f}",
                    officer_count,
                    len(df_merged)
                ]
            }
            
            if self.officer_targets:
                total_target = sum(self.officer_targets.values())
                overall_achievement = (total_collected / total_target * 100) if total_target > 0 else 0
                
                stats_data['Metric'].extend(['Total Target', 'Overall Achievement %', 'Total Remaining'])
                stats_data['Value'].extend([
                    f"{total_target:,.2f}",
                    f"{overall_achievement:.1f}%" if total_target > 0 else "N/A",
                    f"{total_target - total_collected:,.2f}" if total_target > 0 else "N/A"
                ])
            
            stats_df = pd.DataFrame(stats_data)
            stats_df.to_excel(writer, sheet_name='Statistics', index=False)
            
            # Auto-adjust column widths for all sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Reset buffer position
        output.seek(0)
        
        # Load workbook for formatting
        wb = load_workbook(output)
        wb = self.format_excel_report(wb)
        
        # Save formatted workbook to buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def generate_json_report(self, df_collected: pd.DataFrame, 
                           df_merged: pd.DataFrame, 
                           final_df: pd.DataFrame) -> Dict:
        """Generate JSON report for API response."""
        total_collected = df_collected['Collected'].sum() if not df_collected.empty else 0
        officer_count = len(df_collected['SalesRep'].unique()) if not df_collected.empty else 0
        
        report = {
            'status': 'success',
            'timestamp': datetime.now().isoformat(),
            'summary': {
                'total_collected': float(total_collected),
                'total_loans_collected': len(df_collected),
                'officer_count': officer_count,
                'total_loans_processed': len(df_merged),
                'average_collection': float(df_collected['Collected'].mean() if not df_collected.empty else 0)
            }
        }
        
        # Add officer targets if available
        if self.officer_targets:
            total_target = sum(self.officer_targets.values())
            overall_achievement = (total_collected / total_target * 100) if total_target > 0 else 0
            
            report['summary']['total_target'] = float(total_target)
            report['summary']['overall_achievement_percentage'] = float(overall_achievement)
            report['summary']['remaining_target'] = float(total_target - total_collected)
            report['officer_targets'] = self.officer_targets
        
        # Add bucket distribution
        if not df_collected.empty:
            bucket_dist = df_collected.groupby('Bucket')['Collected'].sum().to_dict()
            report['bucket_distribution'] = {k: float(v) for k, v in bucket_dist.items()}
        
        # Add officer performance
        if not df_collected.empty:
            officer_performance = df_collected.groupby('SalesRep').agg({
                'Collected': 'sum',
                'LoanId': 'count'
            }).rename(columns={'LoanId': 'LoansCollected'}).to_dict('index')
            
            # Convert to float for JSON serialization
            officer_performance_clean = {}
            for officer, data in officer_performance.items():
                officer_performance_clean[officer] = {
                    'collected': float(data['Collected']),
                    'loans_collected': int(data['LoansCollected'])
                }
            
            report['officer_performance'] = officer_performance_clean
        
        # Add sample data
        report['sample_data'] = {
            'collected_loans': df_collected.head(10).to_dict(orient='records'),
            'summary_table': final_df.reset_index().to_dict(orient='records')
        }
        
        return report
    
    def process(self, sod_content: bytes, sod_filename: str,
                cur_content: bytes, cur_filename: str,
                officer_targets: Optional[Dict[str, float]] = None,
                output_format: str = 'json') -> Dict:
        """
        Main processing method for API use.
        
        Args:
            sod_content: Bytes content of SOD file
            sod_filename: Name of SOD file
            cur_content: Bytes content of Current file
            cur_filename: Name of Current file
            officer_targets: Optional dictionary of officer targets
            output_format: 'json' or 'excel'
        
        Returns:
            Dictionary with results and either JSON data or Excel bytes
        """
        try:
            # Process data
            df_collected, df_merged, officers = self.process_data(
                sod_content, sod_filename, cur_content, cur_filename
            )
            
            if df_collected.empty:
                return {
                    'status': 'success',
                    'message': 'No collections found in the data',
                    'timestamp': datetime.now().isoformat(),
                    'data': {
                        'total_collected': 0,
                        'officer_count': len(officers),
                        'collected_loans': 0
                    }
                }
            
            # Create formatted table with targets
            final_df = self.create_formatted_table(df_collected, officers, officer_targets)
            
            if output_format.lower() == 'excel':
                # Generate Excel report
                excel_bytes = self.create_excel_report(df_collected, df_merged, final_df)
                
                return {
                    'status': 'success',
                    'message': 'Excel report generated successfully',
                    'timestamp': datetime.now().isoformat(),
                    'format': 'excel',
                    'excel_data': excel_bytes.getvalue(),
                    'filename': f'Arrears_Collection_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                }
            else:
                # Generate JSON report
                json_report = self.generate_json_report(df_collected, df_merged, final_df)
                
                return json_report
            
        except Exception as e:
            return {
                'status': 'error',
                'message': str(e),
                'timestamp': datetime.now().isoformat()
            }

# Command line usage example
if __name__ == "__main__":
    # Example of how to use the processor directly
    import sys
    import json
    
    if len(sys.argv) > 2:
        # Read files from command line
        sod_path = sys.argv[1]
        cur_path = sys.argv[2]
        
        with open(sod_path, 'rb') as f:
            sod_content = f.read()
        
        with open(cur_path, 'rb') as f:
            cur_content = f.read()
        
        processor = ArrearsProcessorAPI()
        result = processor.process(
            sod_content=sod_content,
            sod_filename=os.path.basename(sod_path),
            cur_content=cur_content,
            cur_filename=os.path.basename(cur_path),
            output_format='json'
        )
        
        print(json.dumps(result, indent=2))
    else:
        print("Usage: python Arreas_collected.py <sod_file> <cur_file>")