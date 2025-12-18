# mtd_parameters.py - Refactored for API without tkinter

import pandas as pd
import numpy as np
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
import io
import os
import tempfile
from typing import Dict, Tuple, Optional, Union

warnings.filterwarnings('ignore')

class MTDParametersAPI:
    """API version of MTD Parameters Branch Comparison without tkinter"""
    
    def __init__(self):
        # Initialize data containers
        self.income_data = None
        self.cr_data = None
        self.disb_data = None
        self.merged_data = None
        
    def load_data(self, income_file: Union[str, io.BytesIO], 
                  cr_file: Union[str, io.BytesIO], 
                  disb_file: Union[str, io.BytesIO]) -> Dict:
        """Load data from files (can be file paths or BytesIO objects)"""
        
        try:
            # Load Income Data
            if isinstance(income_file, io.BytesIO):
                self.income_data = pd.read_csv(income_file)
            else:
                self.income_data = pd.read_csv(income_file)
            
            # Clean and validate Income data
            if 'Income (KES)' not in self.income_data.columns:
                # Try to find income column
                income_cols = [col for col in self.income_data.columns if 'income' in col.lower() or 'kes' in col.lower()]
                if income_cols:
                    self.income_data = self.income_data.rename(columns={income_cols[0]: 'Income (KES)'})
                else:
                    raise ValueError("Income column not found in Income file")
            
            self.income_data['Income (KES)'] = pd.to_numeric(
                self.income_data['Income (KES)'].astype(str).str.replace(',', ''), 
                errors='coerce'
            ).fillna(0)
            
            # Load CR Data
            if isinstance(cr_file, io.BytesIO):
                self.cr_data = pd.read_csv(cr_file)
            else:
                self.cr_data = pd.read_csv(cr_file)
            
            # Clean and validate CR data
            cr_columns = ['Collected', 'Uncollected', 'CR %']
            for col in cr_columns:
                if col in self.cr_data.columns:
                    self.cr_data[col] = pd.to_numeric(
                        self.cr_data[col].astype(str).str.replace(',', ''), 
                        errors='coerce'
                    ).fillna(0)
            
            # Load Disbursement Data
            if isinstance(disb_file, io.BytesIO):
                self.disb_data = pd.read_csv(disb_file)
            else:
                self.disb_data = pd.read_csv(disb_file)
            
            # Clean and validate Disbursement data
            if 'Disbursement' in self.disb_data.columns:
                self.disb_data['Disbursement'] = pd.to_numeric(
                    self.disb_data['Disbursement'].astype(str).str.replace(',', ''), 
                    errors='coerce'
                ).fillna(0)
            
            if 'Loan Count' in self.disb_data.columns:
                self.disb_data['Loan Count'] = pd.to_numeric(
                    self.disb_data['Loan Count'].astype(str).str.replace(',', ''), 
                    errors='coerce'
                ).fillna(0).astype(int)
            
            return {
                'status': 'success',
                'message': f'Data loaded successfully! Income: {len(self.income_data)}, CR: {len(self.cr_data)}, Disb: {len(self.disb_data)}',
                'counts': {
                    'income_records': len(self.income_data),
                    'cr_records': len(self.cr_data),
                    'disb_records': len(self.disb_data)
                }
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Error loading data: {str(e)}',
                'error': str(e)
            }
    
    def analyze_data(self, sort_option: str = 'cr_desc') -> Dict:
        """Analyze and merge the loaded data"""
        
        if not all([self.income_data is not None, self.cr_data is not None, self.disb_data is not None]):
            return {
                'status': 'error',
                'message': 'Please load all data files first!'
            }
        
        try:
            # Clean column names
            self.income_data.columns = self.income_data.columns.str.strip()
            self.cr_data.columns = self.cr_data.columns.str.strip()
            self.disb_data.columns = self.disb_data.columns.str.strip()
            
            # Standardize Branch Name column
            for df_name, df in [('income', self.income_data), ('cr', self.cr_data), ('disb', self.disb_data)]:
                # Find branch name column
                branch_cols = [col for col in df.columns if 'branch' in col.lower() or 'name' in col.lower()]
                if branch_cols:
                    df.rename(columns={branch_cols[0]: 'Branch Name'}, inplace=True)
                elif len(df.columns) > 1:
                    # Assume second column is branch name if not found
                    df['Branch Name'] = df.iloc[:, 1]
                else:
                    df['Branch Name'] = f"Branch_{df_name}"
            
            # Clean branch names
            for df in [self.income_data, self.cr_data, self.disb_data]:
                df['Branch Name'] = df['Branch Name'].astype(str).str.strip().str.title()
            
            # Merge data
            merged = pd.merge(self.income_data[['Branch Name', 'Income (KES)']], 
                            self.cr_data[['Branch Name', 'CR %', 'Collected', 'Uncollected']], 
                            on='Branch Name', 
                            how='outer')
            
            merged = pd.merge(merged,
                            self.disb_data[['Branch Name', 'Disbursement', 'Loan Count']],
                            on='Branch Name',
                            how='outer')
            
            # Calculate additional metrics
            merged['Avg Loan Size'] = merged.apply(
                lambda row: row['Disbursement'] / row['Loan Count'] if row['Loan Count'] > 0 else 0,
                axis=1
            ).round(2)
            
            # Calculate weighted Performance Score
            income_norm = merged['Income (KES)'] / merged['Income (KES)'].max() if merged['Income (KES)'].max() > 0 else 0
            cr_norm = merged['CR %'] / 100
            disb_norm = merged['Disbursement'] / merged['Disbursement'].max() if merged['Disbursement'].max() > 0 else 0
            
            merged['Performance Score'] = (
                cr_norm * 0.4 +      # CR weight: 40%
                income_norm * 0.3 +    # Income weight: 30%
                disb_norm * 0.3        # Disbursement weight: 30%
            ) * 100
            
            merged['Performance Score'] = merged['Performance Score'].round(2)
            
            # Handle missing values
            merged = merged.fillna({
                'Income (KES)': 0,
                'CR %': 0,
                'Collected': 0,
                'Uncollected': 0,
                'Disbursement': 0,
                'Loan Count': 0,
                'Avg Loan Size': 0,
                'Performance Score': 0
            })
            
            # Sort data
            sort_map = {
                'cr_desc': ('CR %', False),
                'cr_asc': ('CR %', True),
                'income_desc': ('Income (KES)', False),
                'income_asc': ('Income (KES)', True),
                'disb_desc': ('Disbursement', False),
                'disb_asc': ('Disbursement', True),
                'loans_desc': ('Loan Count', False),
                'loans_asc': ('Loan Count', True),
                'name_asc': ('Branch Name', True),
                'name_desc': ('Branch Name', False),
                'score_desc': ('Performance Score', False)
            }
            
            sort_column, ascending = sort_map.get(sort_option, ('Performance Score', False))
            sorted_data = merged.sort_values(sort_column, ascending=ascending)
            
            # Add rank column
            sorted_data = sorted_data.reset_index(drop=True)
            sorted_data.insert(0, 'Rank', range(1, len(sorted_data) + 1))
            
            self.merged_data = sorted_data
            
            # Calculate summary statistics
            summary_stats = {
                'total_branches': len(sorted_data),
                'total_income': sorted_data['Income (KES)'].sum(),
                'avg_income': sorted_data['Income (KES)'].mean(),
                'avg_cr': sorted_data['CR %'].mean(),
                'total_disbursement': sorted_data['Disbursement'].sum(),
                'avg_disbursement': sorted_data['Disbursement'].mean(),
                'total_loans': sorted_data['Loan Count'].sum(),
                'avg_loan_size': sorted_data['Avg Loan Size'].mean(),
                'avg_performance_score': sorted_data['Performance Score'].mean()
            }
            
            # Get top performers
            top_cr = sorted_data.nlargest(5, 'CR %')[['Branch Name', 'CR %']].to_dict('records')
            top_income = sorted_data.nlargest(5, 'Income (KES)')[['Branch Name', 'Income (KES)']].to_dict('records')
            
            return {
                'status': 'success',
                'message': f'Analysis complete! Processed {len(sorted_data)} branches',
                'data': sorted_data.to_dict('records'),
                'summary': summary_stats,
                'top_performers': {
                    'by_cr': top_cr,
                    'by_income': top_income
                },
                'metadata': {
                    'sort_option': sort_option,
                    'analysis_date': datetime.now().isoformat()
                }
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Error during analysis: {str(e)}',
                'error': str(e)
            }
    
    def export_to_excel(self, output_path: Optional[str] = None) -> Dict:
        """Export analysis to Excel file with YELLOW headers"""
        
        if self.merged_data is None or self.merged_data.empty:
            return {
                'status': 'error',
                'message': 'No data to export! Please analyze data first.'
            }
        
        try:
            # Create a temporary file if no output path provided
            if output_path is None:
                temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                output_path = temp_file.name
                temp_file.close()
            
            # Create workbook with premium styling
            wb = Workbook()
            ws = wb.active
            ws.title = "Branch Performance"
            
            # ------------------------------------------------------------------------
            # UPDATED STYLES FOR YELLOW HEADER TABLE LOOK
            # ------------------------------------------------------------------------
            
            # Main Title Style (Row 1) - Black Background, Yellow Text
            title_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            title_font = Font(bold=True, color="FFFF00", size=14, name='Calibri')
            
            # Column Headers Style (Row 3) - BRIGHT YELLOW Background, Black Text
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_font = Font(bold=True, color="000000", size=11, name='Calibri')
            
            # Subheader/Info Style (Row 2, Summary)
            subheader_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Light yellow
            subheader_font = Font(bold=True, color="000000", size=10, name='Calibri')
            
            # Data Rows
            even_row_fill = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
            
            # Sharp Black Borders for Table Look
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            align_center = Alignment(horizontal="center", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            
            # Prepare data for export
            export_data = self.merged_data.copy()
            
            # Write main header
            ws.merge_cells('A1:J1')
            ws['A1'] = "PREMIUM BRANCH PERFORMANCE REPORT"
            ws['A1'].fill = title_fill
            ws['A1'].font = title_font
            ws['A1'].alignment = align_center
            ws.row_dimensions[1].height = 30
            
            # Write sub-header
            ws.merge_cells('A2:J2')
            ws['A2'] = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].fill = subheader_fill
            ws['A2'].font = subheader_font
            ws['A2'].alignment = align_center
            
            # Write column headers with YELLOW FILL
            headers = [
                'Rank', 'Branch Name', 'Income (KES)', 'CR %',
                'Collected (KES)', 'Uncollected (KES)', 'Disbursement (KES)',
                'Loan Count', 'Avg Loan Size (KES)', 'Performance Score'
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num, value=header)
                cell.fill = header_fill  # Yellow
                cell.font = header_font  # Black Bold
                cell.alignment = align_center
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(col_num)].width = 18
            
            # Add AutoFilter to create "Excel Table" functionality
            ws.auto_filter.ref = f"A3:J{len(export_data)+3}"
            
            # Write data with alternating row colors
            for row_num, (_, row) in enumerate(export_data.iterrows(), 4):
                ws.cell(row=row_num, column=1, value=int(row['Rank'])).alignment = align_center
                ws.cell(row=row_num, column=2, value=str(row['Branch Name'])).alignment = align_left
                ws.cell(row=row_num, column=3, value=float(row['Income (KES)'])).number_format = '#,##0'
                ws.cell(row=row_num, column=4, value=float(row['CR %'])).number_format = '0.00'
                ws.cell(row=row_num, column=5, value=float(row['Collected'])).number_format = '#,##0'
                ws.cell(row=row_num, column=6, value=float(row['Uncollected'])).number_format = '#,##0'
                ws.cell(row=row_num, column=7, value=float(row['Disbursement'])).number_format = '#,##0'
                ws.cell(row=row_num, column=8, value=int(row['Loan Count'])).number_format = '#,##0'
                ws.cell(row=row_num, column=9, value=float(row['Avg Loan Size'])).number_format = '#,##0.00'
                ws.cell(row=row_num, column=10, value=float(row['Performance Score'])).number_format = '0.0'
                
                # Apply alternating row colors
                if row_num % 2 == 0:
                    for col in range(1, 11):
                        ws.cell(row=row_num, column=col).fill = even_row_fill
                
                # Apply borders and alignment
                for col in range(1, 11):
                    ws.cell(row=row_num, column=col).border = thin_border
                    if col in [3, 4, 5, 6, 7, 8, 9, 10]:
                        ws.cell(row=row_num, column=col).alignment = align_right
            
            # Create Summary sheet
            ws_summary = wb.create_sheet(title="Summary")
            
            # Summary statistics with premium formatting
            summary_data = [
                ["SUMMARY STATISTICS", "VALUE"],
                ["", ""],
                ["Total Branches Analyzed", len(export_data)],
                ["Total Income (KES)", export_data['Income (KES)'].sum()],
                ["Average Income (KES)", export_data['Income (KES)'].mean()],
                ["Total Disbursement (KES)", export_data['Disbursement'].sum()],
                ["Average Disbursement (KES)", export_data['Disbursement'].mean()],
                ["Total Loan Count", export_data['Loan Count'].sum()],
                ["Average Loan Size (KES)", export_data['Avg Loan Size'].mean()],
                ["Average CR %", export_data['CR %'].mean()],
                ["Average Performance Score", export_data['Performance Score'].mean()],
                ["", ""],
                ["TOP 5 BRANCHES BY CR %", ""]
            ]
            
            top_cr = export_data.nlargest(5, 'CR %')[['Branch Name', 'CR %']]
            for _, row in top_cr.iterrows():
                summary_data.append([row['Branch Name'], f"{row['CR %']:.2f}%"])
            
            summary_data.append(["", ""])
            summary_data.append(["TOP 5 BRANCHES BY INCOME", ""])
            
            top_income = export_data.nlargest(5, 'Income (KES)')[['Branch Name', 'Income (KES)']]
            for _, row in top_income.iterrows():
                summary_data.append([row['Branch Name'], f"KES {row['Income (KES)']:,.0f}"])
            
            # Write summary
            for row_num, row_data in enumerate(summary_data, 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws_summary.cell(row=row_num, column=col_num, value=value)
                    
                    # Style headers
                    if row_num == 1:
                        cell.fill = title_fill
                        cell.font = title_font
                        cell.alignment = align_center
                    elif row_num <= 12 and col_num == 1:
                        cell.font = Font(bold=True, size=11)
                    elif row_num in [13, 17]:
                        cell.fill = header_fill  # Yellow Headers for sections
                        cell.font = header_font
                        cell.alignment = align_center
                    
                    cell.border = thin_border
            
            # Auto-adjust column widths
            for sheet in [ws, ws_summary]:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 3, 40)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(output_path)
            
            return {
                'status': 'success',
                'message': f'Report exported successfully',
                'file_path': output_path,
                'file_name': os.path.basename(output_path),
                'file_size': os.path.getsize(output_path)
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Error exporting to Excel: {str(e)}',
                'error': str(e)
            }
    
    def generate_charts(self) -> Dict:
        """Generate visualization charts as base64 images"""
        
        if self.merged_data is None or self.merged_data.empty:
            return {
                'status': 'error',
                'message': 'No data to visualize! Please analyze data first.'
            }
        
        try:
            import base64
            from io import BytesIO
            
            charts = {}
            data = self.merged_data.nlargest(20, 'Performance Score')
            
            # Chart 1: Top Performers by Score
            fig1, ax1 = plt.subplots(figsize=(10, 6))
            ax1.barh(data['Branch Name'][:10], data['Performance Score'][:10], 
                    color=plt.cm.viridis(np.linspace(0.2, 0.8, 10)))
            ax1.set_xlabel('Performance Score')
            ax1.set_title('Top 10 Branches by Performance', fontsize=12, fontweight='bold')
            ax1.invert_yaxis()
            ax1.grid(True, alpha=0.3, axis='x')
            plt.tight_layout()
            
            # Save to base64
            buf1 = BytesIO()
            fig1.savefig(buf1, format='png', dpi=100, bbox_inches='tight')
            buf1.seek(0)
            charts['top_performers'] = base64.b64encode(buf1.read()).decode('utf-8')
            plt.close(fig1)
            
            # Chart 2: Income vs CR % Scatter
            fig2, ax2 = plt.subplots(figsize=(10, 6))
            scatter = ax2.scatter(self.merged_data['Income (KES)'], 
                                self.merged_data['CR %'],
                                c=self.merged_data['Performance Score'],
                                cmap='viridis',
                                s=100,
                                alpha=0.6,
                                edgecolors='black',
                                linewidth=0.5)
            ax2.set_xlabel('Income (KES)')
            ax2.set_ylabel('CR %')
            ax2.set_title('Income vs CR % Correlation', fontsize=12, fontweight='bold')
            ax2.grid(True, alpha=0.3)
            plt.colorbar(scatter, ax=ax2, label='Performance Score')
            plt.tight_layout()
            
            buf2 = BytesIO()
            fig2.savefig(buf2, format='png', dpi=100, bbox_inches='tight')
            buf2.seek(0)
            charts['income_vs_cr'] = base64.b64encode(buf2.read()).decode('utf-8')
            plt.close(fig2)
            
            # Chart 3: CR Distribution
            fig3, ax3 = plt.subplots(figsize=(10, 6))
            ax3.hist(self.merged_data['CR %'], bins=15, color='skyblue', edgecolor='black', alpha=0.7)
            ax3.set_xlabel('CR %')
            ax3.set_ylabel('Number of Branches')
            ax3.set_title('CR % Distribution', fontsize=12, fontweight='bold')
            ax3.grid(True, alpha=0.3)
            plt.tight_layout()
            
            buf3 = BytesIO()
            fig3.savefig(buf3, format='png', dpi=100, bbox_inches='tight')
            buf3.seek(0)
            charts['cr_distribution'] = base64.b64encode(buf3.read()).decode('utf-8')
            plt.close(fig3)
            
            return {
                'status': 'success',
                'message': 'Charts generated successfully',
                'charts': charts
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Error generating charts: {str(e)}',
                'error': str(e)
            }
    
    def get_summary_stats(self) -> Dict:
        """Get summary statistics for the analyzed data"""
        
        if self.merged_data is None or self.merged_data.empty:
            return {
                'status': 'error',
                'message': 'No data available! Please analyze data first.'
            }
        
        try:
            data = self.merged_data
            
            summary = {
                'total_branches': len(data),
                'total_income': data['Income (KES)'].sum(),
                'average_income': data['Income (KES)'].mean(),
                'total_disbursement': data['Disbursement'].sum(),
                'average_disbursement': data['Disbursement'].mean(),
                'total_loans': int(data['Loan Count'].sum()),
                'average_cr': data['CR %'].mean(),
                'average_performance_score': data['Performance Score'].mean(),
                'top_performing_branch': {
                    'name': data.loc[data['Performance Score'].idxmax(), 'Branch Name'],
                    'score': float(data['Performance Score'].max())
                },
                'highest_cr_branch': {
                    'name': data.loc[data['CR %'].idxmax(), 'Branch Name'],
                    'cr': float(data['CR %'].max())
                }
            }
            
            return {
                'status': 'success',
                'message': 'Summary statistics retrieved',
                'summary': summary
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Error calculating summary: {str(e)}',
                'error': str(e)
            }


# Helper function for API usage
def process_mtd_parameters(income_file, cr_file, disb_file, sort_option='cr_desc', return_excel=True):
    """
    Main function to process MTD parameters
    
    Args:
        income_file: Path to income CSV file or BytesIO object
        cr_file: Path to CR CSV file or BytesIO object  
        disb_file: Path to disbursement CSV file or BytesIO object
        sort_option: Sorting option (default: 'cr_desc')
        return_excel: Whether to return Excel file path (default: True)
    
    Returns:
        Dictionary with results
    """
    
    analyzer = MTDParametersAPI()
    
    # Load data
    load_result = analyzer.load_data(income_file, cr_file, disb_file)
    if load_result['status'] == 'error':
        return load_result
    
    # Analyze data
    analysis_result = analyzer.analyze_data(sort_option)
    if analysis_result['status'] == 'error':
        return analysis_result
    
    # Get summary
    summary_result = analyzer.get_summary_stats()
    
    # Generate charts
    charts_result = analyzer.generate_charts()
    
    # Export to Excel if requested
    excel_result = None
    if return_excel:
        excel_result = analyzer.export_to_excel()
    
    # Combine results
    result = {
        'status': 'success',
        'load_result': load_result,
        'analysis_result': analysis_result,
        'summary_result': summary_result,
        'charts_result': charts_result
    }
    
    if excel_result and excel_result['status'] == 'success':
        result['excel_result'] = excel_result
    
    return result


# For direct script execution
if __name__ == "__main__":
    # Example usage
    import sys
    
    if len(sys.argv) < 4:
        print("Usage: python mtd_parameters.py <income_file> <cr_file> <disb_file> [sort_option]")
        sys.exit(1)
    
    income_file = sys.argv[1]
    cr_file = sys.argv[2]
    disb_file = sys.argv[3]
    sort_option = sys.argv[4] if len(sys.argv) > 4 else 'cr_desc'
    
    result = process_mtd_parameters(income_file, cr_file, disb_file, sort_option)
    
    if result['status'] == 'success':
        print("✓ Analysis completed successfully!")
        print(f"Processed {result['analysis_result']['data']['total_branches']} branches")
        
        if 'excel_result' in result:
            print(f"Excel file saved: {result['excel_result']['file_path']}")
    else:
        print(f"✗ Error: {result.get('message', 'Unknown error')}")