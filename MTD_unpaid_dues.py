# arrears_risk_analysis.py - Refactored for API without tkinter

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
import numpy as np
import os
import sys
import io
import tempfile
from typing import Dict, Optional, Union
import warnings

warnings.filterwarnings('ignore')

# Set matplotlib backend to non-interactive for server use
plt.switch_backend('Agg')

# =========================================================
# CONFIGURATION
# =========================================================
CHART_DPI = 200
plt.style.use('ggplot')

class ArrearsRiskAnalyzer:
    """API version of Arrears Risk Analysis without tkinter"""
    
    def __init__(self):
        self.df = None
        self.customer_risk = None
        self.output_file = None
        self.temp_files = []
    
    def load_and_clean_data(self, file_input: Union[str, io.BytesIO]) -> Dict:
        """Loads Excel/CSV, standardizes columns, and cleans data types."""
        try:
            if isinstance(file_input, io.BytesIO):
                # Reset stream position
                file_input.seek(0)
                
                # Try to detect file type by extension or content
                try:
                    # Try Excel first
                    df = pd.read_excel(file_input)
                except:
                    # Reset and try CSV
                    file_input.seek(0)
                    df = pd.read_csv(file_input)
            else:
                # File path
                if not os.path.exists(file_input):
                    return {
                        'status': 'error',
                        'message': f"The file '{file_input}' was not found."
                    }
                
                if file_input.lower().endswith('.csv'):
                    df = pd.read_csv(file_input)
                else:
                    df = pd.read_excel(file_input)
            
            # Standardize headers
            df.columns = [str(c).strip() for c in df.columns]
            
            # --- MAPPING LOGIC ---
            colmap = {
                "FullNames": "FullNames",
                "PhoneNumber": "PhoneNumber",
                "InstallmentNo": "InstallmentNo",
                "Amount Due": "AmountDue",
                "Arrears": "Arrears",
                "LoanBalance": "LoanBalance",
                "FieldOfficer": "FieldOfficer",
                "Principal": "FundedAmount",
            }
            df = df.rename(columns=colmap)

            # Handle missing essential columns
            if "FundedAmount" not in df.columns:
                df["FundedAmount"] = 0
            if "FieldOfficer" not in df.columns:
                df["FieldOfficer"] = "Unassigned"

            # Numeric conversion
            numeric_cols = ["InstallmentNo", "AmountDue", "Arrears", "LoanBalance", "FundedAmount"]
            for c in numeric_cols:
                if c in df.columns:
                    df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

            # Robust Phone Cleaning
            if "PhoneNumber" in df.columns:
                df["PhoneNumber"] = df["PhoneNumber"].astype(str).str.replace(r"\D", "", regex=True)
            
            self.df = df
            
            return {
                'status': 'success',
                'message': f'Successfully loaded {len(df)} records',
                'record_count': len(df),
                'columns': list(df.columns)
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Error reading file: {str(e)}'
            }
    
    def calculate_risk_metrics(self) -> Dict:
        """Calculates risk scores and categories using Original Logic."""
        if self.df is None:
            return {
                'status': 'error',
                'message': 'No data loaded. Please load data first.'
            }
        
        try:
            df = self.df
            
            # Sort to identify latest status
            df_sorted = df.sort_values(by=["PhoneNumber", "InstallmentNo"])
            latest_status = df_sorted.groupby("PhoneNumber").tail(1).copy()

            # Calculate Total Missed Installments
            df["MissedFlag"] = (df["Arrears"] > 0).astype(int)
            missed_inst = (
                df.groupby("PhoneNumber")["MissedFlag"]
                .sum()
                .reset_index()
                .rename(columns={"MissedFlag": "MissedInstallments"})
            )

            # Merge
            customer_risk = latest_status.merge(missed_inst, on="PhoneNumber", how="left")
            customer_risk["MissedInstallments"] = customer_risk["MissedInstallments"].fillna(0).astype(int)

            # Risk Scoring Formula (Original Weights)
            arrears_cap = customer_risk["Arrears"].quantile(0.95)
            balance_cap = customer_risk["LoanBalance"].quantile(0.95)
            
            arrears_capped = np.minimum(customer_risk["Arrears"], arrears_cap)
            balance_capped = np.minimum(customer_risk["LoanBalance"], balance_cap)
            missed_capped = np.minimum(customer_risk["MissedInstallments"], 12)

            customer_risk["RiskScore"] = (
                arrears_capped * 0.50 +
                missed_capped * 0.35 +
                balance_capped * 0.15
            )

            # Risk Categorization (Original Thresholds)
            q40 = customer_risk["RiskScore"].quantile(0.40)
            q75 = customer_risk["RiskScore"].quantile(0.75)

            def determine_category(row):
                # 1. HARD RULE: If missed > 4 installments, High Risk
                if row["MissedInstallments"] >= 4: return "High Risk"
                # 2. HARD RULE: Significant Arrears (> 5000) is High Risk
                if row["Arrears"] > 5000: return "High Risk"
                
                # 3. Percentile Rules
                score = row["RiskScore"]
                if score >= q75: return "High Risk"
                elif score >= q40: return "Medium Risk"
                else: return "Low Risk"

            customer_risk["RiskCategory"] = customer_risk.apply(determine_category, axis=1)
            customer_risk = customer_risk.sort_values(by="RiskScore", ascending=False).reset_index(drop=True)
            
            self.customer_risk = customer_risk
            
            # Calculate summary statistics
            risk_counts = customer_risk["RiskCategory"].value_counts()
            summary = {
                'total_customers': len(customer_risk),
                'high_risk': int(risk_counts.get('High Risk', 0)),
                'medium_risk': int(risk_counts.get('Medium Risk', 0)),
                'low_risk': int(risk_counts.get('Low Risk', 0)),
                'total_arrears': float(customer_risk["Arrears"].sum()),
                'avg_arrears': float(customer_risk["Arrears"].mean()),
                'customers_in_arrears': int((customer_risk["Arrears"] > 0).sum()),
                'loans_in_arrears_percent': float(((customer_risk["Arrears"] > 0).mean() * 100))
            }
            
            return {
                'status': 'success',
                'message': f'Risk analysis complete. Processed {len(customer_risk)} customers',
                'summary': summary,
                'risk_distribution': risk_counts.to_dict()
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Error calculating risk metrics: {str(e)}'
            }
    
    def build_printable_risk_sheet(self):
        """Build printable risk sheet with officer blocks"""
        if self.customer_risk is None:
            return pd.DataFrame()
        
        customer_risk = self.customer_risk.copy()
        
        # --- UPDATED: EXCLUDE COLUMNS LOGIC ---
        exclude_patterns = [
            "loanid", "loan id", 
            "amountpaid", "amount paid", 
            "loanamount", "loan amount", 
            "productname", "product name", 
            "unittitle", "unit title", 
            "duedate", "due date", 
            "fundedamount", "funded amount"
        ]
        
        cols_to_drop = []
        for col in customer_risk.columns:
            if col.lower() in exclude_patterns:
                cols_to_drop.append(col)
                
        if cols_to_drop:
            customer_risk = customer_risk.drop(columns=cols_to_drop)
        # --------------------------------------

        officer_blocks_list = []
        
        def make_spacer():
            return pd.DataFrame([{c: None for c in customer_risk.columns}])

        def make_row(officer, name_val, arrears=None, balance=None, missed=None, score=None, cat=None):
            data = {c: None for c in customer_risk.columns}
            data["FieldOfficer"] = officer
            data["FullNames"] = name_val
            data["Arrears"] = arrears
            data["LoanBalance"] = balance
            data["MissedInstallments"] = missed
            data["RiskScore"] = score
            data["RiskCategory"] = cat
            return pd.DataFrame([data])

        for officer_name, officer_df in customer_risk.groupby("FieldOfficer"):
            officer_blocks_list.append(make_row(officer_name, f"{officer_name} PORTFOLIO"))

            for cat in ["High Risk", "Medium Risk", "Low Risk"]:
                block = officer_df[officer_df["RiskCategory"] == cat].copy()
                if block.empty: continue
                
                block = block.sort_values("RiskScore", ascending=False)
                officer_blocks_list.append(block)
                
                officer_blocks_list.append(make_row(
                    officer_name, f"{cat} SUBTOTAL", 
                    arrears=block["Arrears"].sum(),
                    balance=block["LoanBalance"].sum(),
                    missed=block["MissedInstallments"].sum(),
                    score=block["RiskScore"].sum(),
                    cat=cat
                ))
            officer_blocks_list.append(make_spacer())

        if not officer_blocks_list: 
            return pd.DataFrame()
        
        return pd.concat(officer_blocks_list, ignore_index=True)
    
    def build_early_arrears_report(self):
        """Build early arrears report"""
        if self.customer_risk is None or self.df is None:
            return pd.DataFrame()
        
        customer_risk = self.customer_risk.copy()
        
        early_mask = (customer_risk["InstallmentNo"] <= 2) & (customer_risk["Arrears"] > 0)
        early_df = customer_risk.loc[early_mask].copy()
        early_df = early_df.sort_values(by=["FieldOfficer", "Arrears"], ascending=[True, False])
        
        # --- UPDATED: Removed 'FundedAmount' from list ---
        cols = [
            "FieldOfficer", 
            "FullNames", 
            "PhoneNumber", 
            "InstallmentNo", 
            "AmountDue", 
            "Arrears", 
            "LoanBalance", 
            "RiskCategory", 
            "DiagnosticNote"
        ]
        
        blocks = []
        def create_row(data_dict): return pd.DataFrame([data_dict])

        for officer, block in early_df.groupby("FieldOfficer"):
            header = {c: None for c in cols}
            header["FieldOfficer"] = officer
            header["DiagnosticNote"] = f"{officer} - Early Arrears"
            blocks.append(create_row(header))
            
            block_data = block.copy()
            block_data["DiagnosticNote"] = "New Loan Arrears"
            blocks.append(block_data[cols])
            
            sub = {c: None for c in cols}
            sub["FieldOfficer"] = officer
            sub["FullNames"] = "SUBTOTAL"
            sub["AmountDue"] = block["AmountDue"].sum()
            sub["Arrears"] = block["Arrears"].sum()
            sub["LoanBalance"] = block["LoanBalance"].sum()
            blocks.append(create_row(sub))
            blocks.append(create_row({c: None for c in cols}))
            
        if not blocks: 
            return pd.DataFrame(columns=cols)
        
        grand_total = {c: None for c in cols}
        grand_total["FieldOfficer"] = "TOTAL"
        grand_total["FullNames"] = "GRAND TOTAL"
        grand_total["AmountDue"] = early_df["AmountDue"].sum()
        grand_total["Arrears"] = early_df["Arrears"].sum()
        grand_total["LoanBalance"] = early_df["LoanBalance"].sum()
        blocks.append(create_row(grand_total))
        
        return pd.concat(blocks, ignore_index=True)
    
    def generate_summary_statistics(self) -> Dict:
        """Generate summary statistics and reports"""
        if self.customer_risk is None:
            return {}
        
        customer_risk = self.customer_risk
        
        # Officer arrears summary
        officer_arrears = customer_risk.groupby("FieldOfficer")["Arrears"].sum().reset_index().sort_values("Arrears", ascending=False)
        
        # Portfolio summary
        summary_dict = {
            "Total Customers": customer_risk["PhoneNumber"].nunique(),
            "Customers in Arrears": int((customer_risk["Arrears"] > 0).sum()),
            "Portfolio Arrears (Sum)": float(customer_risk["Arrears"].sum()),
            "Average Arrears per Customer": round(float(customer_risk["Arrears"].mean()), 2),
            "Loans in Arrears %": round(((customer_risk["Arrears"] > 0).mean() * 100.0), 2),
            "High Risk Customers": int((customer_risk["RiskCategory"] == "High Risk").sum()),
            "Medium Risk Customers": int((customer_risk["RiskCategory"] == "Medium Risk").sum()),
            "Low Risk Customers": int((customer_risk["RiskCategory"] == "Low Risk").sum()),
        }
        portfolio_summary = pd.DataFrame([summary_dict])
        
        # Officer risk matrix
        officer_matrix = pd.pivot_table(customer_risk, index="FieldOfficer", columns="RiskCategory", values="Arrears", aggfunc="sum", fill_value=0)
        for c in ["High Risk", "Medium Risk", "Low Risk"]:
            if c not in officer_matrix.columns: 
                officer_matrix[c] = 0
        officer_matrix["TotalArrears"] = officer_matrix.sum(axis=1)
        
        officer_matrix["HighRisk%"] = (officer_matrix["High Risk"] / officer_matrix["TotalArrears"] * 100).fillna(0)
        officer_matrix = officer_matrix.sort_values("TotalArrears", ascending=False).reset_index()
        
        # Add Grand Total to Matrix
        grand_total_vals = officer_matrix.sum(numeric_only=True).to_dict()
        grand_total_row = {"FieldOfficer": "TOTAL", **grand_total_vals}
        
        if grand_total_row["TotalArrears"] > 0:
            grand_total_row["HighRisk%"] = (grand_total_row["High Risk"] / grand_total_row["TotalArrears"]) * 100
        else:
            grand_total_row["HighRisk%"] = 0
            
        officer_matrix = pd.concat([officer_matrix, pd.DataFrame([grand_total_row])], ignore_index=True)
        
        # Arrears by installment
        arrears_by_inst = self.df.groupby("InstallmentNo")["Arrears"].max().reset_index().sort_values("InstallmentNo")
        
        return {
            'officer_arrears': officer_arrears,
            'portfolio_summary': portfolio_summary,
            'officer_matrix': officer_matrix,
            'arrears_by_inst': arrears_by_inst,
            'printable_risk': self.build_printable_risk_sheet(),
            'early_arrears': self.build_early_arrears_report()
        }
    
    def generate_charts(self, temp_dir: str) -> Dict:
        """Generate charts and save to temporary directory"""
        if self.customer_risk is None:
            return {}
        
        try:
            customer_risk = self.customer_risk
            summaries = self.generate_summary_statistics()
            
            charts_info = {}
            
            # 1. Total Arrears by Officer
            plt.figure(figsize=(8, 5))
            bars = plt.bar(summaries['officer_arrears']["FieldOfficer"], 
                          summaries['officer_arrears']["Arrears"], 
                          color='#3498db')
            plt.title("Total Arrears by Field Officer", fontsize=12, fontweight='bold')
            plt.xticks(rotation=45, ha="right")
            for bar in bars:
                height = bar.get_height()
                plt.text(bar.get_x() + bar.get_width()/2, height, 
                        f"{int(height):,}", ha='center', va='bottom', fontsize=8)
            plt.tight_layout()
            
            chart1_path = os.path.join(temp_dir, "chart_arrears.png")
            plt.savefig(chart1_path, dpi=CHART_DPI)
            plt.close()
            charts_info['arrears_by_officer'] = chart1_path
            
            # 2. Arrears by Installment
            plt.figure(figsize=(8, 5))
            plt.bar(summaries['arrears_by_inst']["InstallmentNo"], 
                   summaries['arrears_by_inst']["Arrears"], 
                   color='#e74c3c')
            plt.title("Worst Arrears by Loan Age", fontsize=12, fontweight='bold')
            plt.tight_layout()
            
            chart2_path = os.path.join(temp_dir, "chart_age.png")
            plt.savefig(chart2_path, dpi=CHART_DPI)
            plt.close()
            charts_info['arrears_by_installment'] = chart2_path
            
            # 3. Top 10 High Risk
            top10 = customer_risk.sort_values("RiskScore", ascending=False).head(10)
            plt.figure(figsize=(8, 5))
            colors = plt.cm.Reds(np.linspace(0.5, 1, 10))
            plt.barh(top10["FullNames"], top10["RiskScore"], color=colors)
            plt.title("Top 10 Highest Risk Customers", fontsize=12, fontweight='bold')
            plt.gca().invert_yaxis()
            plt.tight_layout()
            
            chart3_path = os.path.join(temp_dir, "chart_top10.png")
            plt.savefig(chart3_path, dpi=CHART_DPI)
            plt.close()
            charts_info['top_10_risk'] = chart3_path
            
            # 4. Pie Chart
            risk_counts = customer_risk["RiskCategory"].value_counts().reindex(["High Risk", "Medium Risk", "Low Risk"]).fillna(0)
            
            # Breakdown Logic
            high_risk_only = customer_risk[customer_risk["RiskCategory"] == "High Risk"]
            high_risk_split = high_risk_only.groupby("FieldOfficer")["Arrears"].sum().sort_values(ascending=False)
            high_total = high_risk_split.sum()
            if high_total > 0:
                breakdown_lines = [f"{off}: {(arr/high_total)*100:.1f}%" for off, arr in high_risk_split.items()]
                breakdown_text = "High Risk split:\n" + "\n".join(breakdown_lines)
            else:
                breakdown_text = ""

            plt.figure(figsize=(7, 6))
            plt.pie(risk_counts, labels=risk_counts.index, autopct="%1.1f%%", 
                   colors=['#e74c3c', '#f1c40f', '#2ecc71'], startangle=90)
            plt.title("Portfolio Risk Composition", fontweight='bold')
            plt.text(-1.5, 1.0, breakdown_text, fontsize=8, va="top")
            plt.tight_layout()
            
            chart4_path = os.path.join(temp_dir, "chart_pie.png")
            plt.savefig(chart4_path, dpi=CHART_DPI)
            plt.close()
            charts_info['risk_composition'] = chart4_path
            
            return charts_info
            
        except Exception as e:
            print(f"Error generating charts: {e}")
            return {}
    
    def format_excel_workbook(self, filepath: str):
        """Applies auto-fit, currency formatting, and styles."""
        wb = load_workbook(filepath)
        currency_fmt = '#,##0.00'
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 1. Auto-fit Columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: 
                        pass
                adjusted_width = (max_length + 2) * 1.1
                if adjusted_width > 50: 
                    adjusted_width = 50 
                ws.column_dimensions[column_letter].width = adjusted_width
                
                # 2. Currency Formatting
                header_val = ws[f"{column_letter}1"].value
                if header_val and any(x in str(header_val) for x in ["Arrears", "Balance", "Amount", "Funded", "Principal"]):
                    for cell in column[1:]: 
                        cell.number_format = currency_fmt

        # 3. Conditional Formatting
        if "Customer Risk Ranking" in wb.sheetnames:
            ws = wb["Customer Risk Ranking"]
            risk_col_idx = None
            for i, cell in enumerate(ws[1], 1):
                if cell.value == "RiskCategory": 
                    risk_col_idx = i
            
            if risk_col_idx:
                col_char = ws.cell(1, risk_col_idx).column_letter
                rng = f"{col_char}2:{col_char}{ws.max_row}"
                ws.conditional_formatting.add(rng, FormulaRule(
                    formula=[f'${col_char}2="High Risk"'], 
                    fill=PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                ))
                ws.conditional_formatting.add(rng, FormulaRule(
                    formula=[f'${col_char}2="Medium Risk"'], 
                    fill=PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                ))
                ws.conditional_formatting.add(rng, FormulaRule(
                    formula=[f'${col_char}2="Low Risk"'], 
                    fill=PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                ))

        # 4. Bold Special Rows
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                txt = " ".join([str(c.value).upper() for c in row[:2] if c.value])
                if any(x in txt for x in ["SUBTOTAL", "TOTAL", "PORTFOLIO", "EARLY ARREARS"]):
                    for cell in row: 
                        cell.font = Font(bold=True)
                        if "SUBTOTAL" in txt or "TOTAL" in txt:
                            cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

        wb.save(filepath)
    
    def insert_charts_into_excel(self, excel_path: str, charts_info: Dict):
        """Insert generated charts into Excel workbook"""
        if not charts_info:
            return
        
        wb = load_workbook(excel_path)
        
        # Insert charts into appropriate sheets
        if "Officer Summary" in wb.sheetnames and 'arrears_by_officer' in charts_info:
            ws = wb["Officer Summary"]
            img = Image(charts_info['arrears_by_officer'])
            ws.add_image(img, "E2")
        
        if "Arrears by Installment" in wb.sheetnames and 'arrears_by_installment' in charts_info:
            ws = wb["Arrears by Installment"]
            img = Image(charts_info['arrears_by_installment'])
            ws.add_image(img, "E2")
        
        if "Customer Risk Ranking" in wb.sheetnames and 'top_10_risk' in charts_info:
            ws = wb["Customer Risk Ranking"]
            img = Image(charts_info['top_10_risk'])
            ws.add_image(img, "J2")
        
        if "Portfolio Summary" in wb.sheetnames and 'risk_composition' in charts_info:
            ws = wb["Portfolio Summary"]
            img = Image(charts_info['risk_composition'])
            ws.add_image(img, "F2")
        
        wb.save(excel_path)
    
    def analyze(self, file_input: Union[str, io.BytesIO], output_path: Optional[str] = None) -> Dict:
        """
        Main analysis function
        
        Args:
            file_input: File path or BytesIO object
            output_path: Optional output path for Excel file
            
        Returns:
            Dictionary with analysis results
        """
        try:
            # Create temporary directory for charts
            temp_dir = tempfile.mkdtemp()
            self.temp_files.append(temp_dir)
            
            # Step 1: Load data
            load_result = self.load_and_clean_data(file_input)
            if load_result['status'] == 'error':
                return load_result
            
            # Step 2: Calculate risk metrics
            risk_result = self.calculate_risk_metrics()
            if risk_result['status'] == 'error':
                return risk_result
            
            # Step 3: Generate summary statistics
            summaries = self.generate_summary_statistics()
            
            # Step 4: Generate charts
            charts_info = self.generate_charts(temp_dir)
            
            # Step 5: Create output Excel file
            if output_path is None:
                # Create temporary output file
                temp_output = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                output_path = temp_output.name
                temp_output.close()
                self.temp_files.append(output_path)
            
            # Write data to Excel
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                summaries['printable_risk'].to_excel(writer, sheet_name="Customer Risk Ranking", index=False)
                summaries['officer_arrears'].to_excel(writer, sheet_name="Officer Summary", index=False)
                summaries['officer_matrix'].to_excel(writer, sheet_name="Officer Risk Breakdown", index=False)
                summaries['portfolio_summary'].to_excel(writer, sheet_name="Portfolio Summary", index=False)
                summaries['arrears_by_inst'].to_excel(writer, sheet_name="Arrears by Installment", index=False)
                summaries['early_arrears'].to_excel(writer, sheet_name="Early Arrears by Officer", index=False)
            
            # Step 6: Insert charts and format Excel
            self.insert_charts_into_excel(output_path, charts_info)
            self.format_excel_workbook(output_path)
            
            # Store output file path
            self.output_file = output_path
            
            # Prepare response
            response = {
                'status': 'success',
                'message': f'Analysis complete. Processed {len(self.df)} records',
                'output_file': output_path,
                'output_filename': os.path.basename(output_path),
                'summary': risk_result['summary'],
                'record_count': load_result['record_count'],
                'risk_distribution': risk_result['risk_distribution'],
                'chart_count': len(charts_info)
            }
            
            return response
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Error during analysis: {str(e)}'
            }
    
    def cleanup(self):
        """Clean up temporary files"""
        import shutil
        import logging
        logger = logging.getLogger(__name__)
        
        for temp_file in self.temp_files:
            try:
                if os.path.isdir(temp_file):
                    shutil.rmtree(temp_file, ignore_errors=True)
                elif os.path.exists(temp_file):
                    os.remove(temp_file)
            except Exception as e:
                logger.warning(f"Failed to cleanup {temp_file}: {e}")


# Helper function for API usage
def analyze_arrears_risk(file_input: Union[str, io.BytesIO], output_path: Optional[str] = None) -> Dict:
    """
    Convenience function for API usage
    
    Args:
        file_input: File path or BytesIO object
        output_path: Optional output path for Excel file
        
    Returns:
        Dictionary with analysis results
    """
    analyzer = ArrearsRiskAnalyzer()
    result = analyzer.analyze(file_input, output_path)
    
    # Clean up temporary files on error
    if result['status'] == 'error':
        analyzer.cleanup()
    
    return result


# For direct script execution
if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Analyze arrears risk from loan data')
    parser.add_argument('input_file', help='Path to input Excel or CSV file')
    parser.add_argument('-o', '--output', help='Output file path (optional)')
    
    args = parser.parse_args()
    
    result = analyze_arrears_risk(args.input_file, args.output)
    
    if result['status'] == 'success':
        print(f"✓ Analysis completed successfully!")
        print(f"Output file: {result['output_file']}")
        print(f"Summary: {result['summary']}")
    else:
        print(f"✗ Error: {result.get('message', 'Unknown error')}")